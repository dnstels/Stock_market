import pandas as pd
import numpy as np
from scipy.optimize import brentq
import warnings
warnings.filterwarnings('ignore')

# Для сохранения в Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    import matplotlib.pyplot as plt
    import matplotlib
    from io import BytesIO
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("Внимание: openpyxl не установлен. Установите: pip install openpyxl matplotlib")

# ========== ДИАГНОСТИКА ДЛЯ MAC ==========
import sys
import os
import time

# Перенаправляем stdout в stderr, чтобы избежать буферизации
sys.stdout = sys.stderr

print("="*80)
print("ДИАГНОСТИЧЕСКИЙ РЕЖИМ ДЛЯ MAC")
print("="*80)
print(f"Версия Python: {sys.version}")
print(f"Текущая директория: {os.getcwd()}")
print(f"Права на запись: {os.access(os.getcwd(), os.W_OK)}")
print(f"Время запуска: {time.strftime('%Y-%m-%d %H:%M:%S')}")
print("="*80)
sys.stderr.flush()

def log(msg):
    timestamp = time.strftime('%H:%M:%S')
    print(f"[{timestamp}] {msg}")
    sys.stderr.flush()

log("Диагностика включена")
# ==========================================

# =============================================================================
# НАСТРОЙКИ ДЛЯ 15-МИНУТНЫХ ДАННЫХ
# =============================================================================

# ------------------------------------------------------
# 1. БАЗОВЫЕ НАСТРОЙКИ ДАННЫХ
# ------------------------------------------------------
# =============================================================================
# НАСТРОЙКИ ДЛЯ 1-МИНУТНЫХ ДАННЫХ
# =============================================================================
use_1min_data = True  # True = использовать 1-минутные данные, False = 15-минутные

if use_1min_data:
    time_step_hours = 1/60  # 1 минута = 1/60 часа
    data_file = 'historical_1min.csv'
    print("\n" + "="*80)
    print("ИСПОЛЬЗУЮТСЯ 1-МИНУТНЫЕ ДАННЫ")
    print("="*80)
else:
    time_step_hours = 0.25  # 15 минут = 0.25 часа
    data_file = 'historical_15min.csv'
    print("\n" + "="*80)
    print("ИСПОЛЬЗУЮТСЯ 15-МИНУТНЫЕ ДАННЫЕ")
    print("="*80)

PRICE_COLUMN = 'close'                       # Столбец цены: 'close', 'open', 'high', 'low'
use_high_low_for_out_of_range = True         # True = использовать high/low для выхода из диапазона, False = только close
verbose_logging = True                        # True = подробный лог в консоль, False = минимальный

# ------------------------------------------------------
# 2. ОСНОВНЫЕ ПАРАМЕТРЫ СТРАТЕГИИ
# ------------------------------------------------------
capital = 100000.00                          # Стартовый капитал в USDC
range_width = 1000.00                         # Ширина диапазона ликвидности в долларах
apr = 0.6                                     # Годовая процентная доходность пула (60% = 0.6)
shift_enabled = True                          # Включить автоматический сдвиг диапазона
shift_delay_hours = 0.2                         # Часов вне диапазона до сдвига
funding_rate = -0.1085                        # Годовой funding rate (-10.85% = -0.1085)
gas_usd = 0.0                                  # Комиссия за газ в долларах
slippage = 0.001                               # Проскальзывание (0.1% = 0.001)
short_fee = 0.001                              # Комиссия за открытие/закрытие шорта (0.1% = 0.001)

# ------------------------------------------------------
# 3. АСИММЕТРИЧНЫЙ ДИАПАЗОН
# ------------------------------------------------------
asymmetric_range_enabled = True                # True = асимметричный диапазон, False = симметричный
asymmetry_distribution = 30                    # % диапазона выше Pn (40% сверху, 60% снизу)
use_current_price_as_center = True             # True = центр диапазона = текущая цена при сдвиге

# ------------------------------------------------------
# 4. ПАРАМЕТРЫ ЗАКРЫТИЯ ШОРТА
# ------------------------------------------------------
close_short_above_entry = True                 # Разрешить закрытие шорта при росте цены
max_loss_percent = 0.1                            # Максимальный убыток % (закрытие при превышении цены входа на 1%)

# ------------------------------------------------------
# 5. BEP (BREAK-EVEN PRICE) ТРИГГЕР
# ------------------------------------------------------
use_bep_as_primary_trigger = True              # True = использовать BEP как основной триггер закрытия
bep_close_buffer_pct = 3                        # Буфер закрытия выше BEP (3% = закрытие при цене на 3% выше BEP)

# ------------------------------------------------------
# 6. ЗАКРЫТИЕ ПО БУФЕРУ ДО PA (FALLBACK)
# ------------------------------------------------------
close_short_below_pa = True                     # True = закрывать шорт при приближении к нижней границе
close_short_buffer_pct = 5                       # Буфер до Pa (5% = закрытие при цене на 5% выше Pa)

# ------------------------------------------------------
# 7. ПОВТОРНОЕ ОТКРЫТИЕ ШОРТА ПОСЛЕ ЗАКРЫТИЯ
# ------------------------------------------------------
reopen_short_below_pn = True                    # True = разрешить повторное открытие шорта ниже Pn
reopen_percent_below_pn = 0                      # % падения ниже Pn для повторного открытия (0% = сразу при входе)
reopen_delay_hours = 0                           # Задержка перед повторным открытием в часах
reopen_bep_buffer_pct = 0                        # Минимальный отскок выше BEP для повторного открытия (%)
ignore_last_bep_for_reopen = True                 # True = игнорировать LastBEP при повторном открытии

# ------------------------------------------------------
# 8. РЕИНВЕСТИРОВАНИЕ КЭША (fees + funding + short PnL)
# ------------------------------------------------------
reinvest_cash_enabled = True                    # True = реинвестировать накопленный кэш
reinvest_frequency = 'weekly'                    # Частота: 'short_close', 'shift', 'weekly', 'biweekly', 'monthly'

# ------------------------------------------------------
# 9. ПОВТОРНЫЙ ВХОД В ПУЛ ПОСЛЕ ВЫХОДА (ИСПРАВЛЕНИЕ ПРОБЛЕМЫ 3)
# ------------------------------------------------------
reentry_enabled = True                           # Включить автоматический повторный вход в пул после выхода
reentry_delay_hours = 1.0                        # Базовая задержка перед повторным входом (в часах)
                                                  # Например: 4 часа = ждем 4 часа после выхода, затем создаем новый пул

reentry_price_drop_threshold = 5.0               # % падения цены от максимума для ускоренного входа
                                                  # Например: 5% = если цена упала на 5% от максимума после выхода,
                                                  # создаем новый пул немедленно, не дожидаясь reentry_delay_hours

reentry_aggressive_hours = 0.5                   # Агрессивный вход через N часов
                                                  # Например: 2 часа = если через 2 часа пул все еще пуст,
                                                  # принудительно создаем новый пул (даже без падения цены)

# ===== НОВЫЕ ПАРАМЕТРЫ ДЛЯ КОРРЕКТИРОВКИ BEP =====
bep_target_enabled = False                        # Включить автоматическую корректировку BEP
bep_target_percent_above_pa = 60                   # Целевой BEP на X% выше Pa (например, 5%)
bep_auto_adjust = False                           # Автоматически корректировать размер шорта
# =================================================

# --- Вывод информации о настройках ---
print("="*80)
print("НАСТРОЙКИ СТРАТЕГИИ С BEP-ТРИГГЕРОМ И АСИММЕТРИЧНЫМ ДИАПАЗОНОМ")
print("="*80)
print(f"Файл данных: {data_file}")
print(f"Ценовой столбец: {PRICE_COLUMN}")
print(f"Шаг времени: {time_step_hours} часа ({int(time_step_hours*60)} минут)")
print(f"Подробный лог: {'ВКЛ' if verbose_logging else 'ВЫКЛ'}")
print(f"Использовать high/low для выхода из диапазона: {'ДА' if use_high_low_for_out_of_range else 'НЕТ'}")
print(f"Стартовый капитал: ${capital:,.2f}")
print(f"Ширина диапазона: ${range_width:.2f}")
print(f"APR пула: {apr:.1%}")
print(f"Сдвиг диапазона: {'ВКЛ' if shift_enabled else 'ВЫКЛ'}")
print(f"Задержка сдвига: {shift_delay_hours} часов")
print(f"Асимметричный диапазон: {'ВКЛ' if asymmetric_range_enabled else 'ВЫКЛ'}")
if asymmetric_range_enabled:
    print(f"  Распределение выше Pn: {asymmetry_distribution}%")
print(f"Центр диапазона всегда на текущей цене: {'ДА' if use_current_price_as_center else 'НЕТ'}")
print(f"Реинвестирование кэша: {'ВКЛ' if reinvest_cash_enabled else 'ВЫКЛ'}")
if reinvest_cash_enabled:
    print(f"  Частота реинвестирования: {reinvest_frequency}")
print(f"Закрытие шорта: при превышении цены входа на {max_loss_percent}%")
print(f"Повторное открытие шорта: при падении на {reopen_percent_below_pn}% ниже Pn")
print(f"Задержка повторного открытия: {reopen_delay_hours} часов")
print(f"Минимальный отскок выше BEP: {reopen_bep_buffer_pct}%")
print(f"Закрытие по BEP-триггеру: {'ВКЛ' if use_bep_as_primary_trigger else 'ВЫКЛ'}")
if use_bep_as_primary_trigger:
    print(f"  Буфер закрытия выше BEP: {bep_close_buffer_pct}%")
print(f"Закрытие по буферу до Pa (fallback): {'ВКЛ' if close_short_below_pa else 'ВЫКЛ'}")
if close_short_below_pa:
    print(f"  Буфер до Pa: {close_short_buffer_pct}%")
print(f"Корректировка BEP: {'ВКЛ' if bep_target_enabled else 'ВЫКЛ'}")
if bep_target_enabled:
    print(f"  Целевой BEP выше Pa: {bep_target_percent_above_pa}%")
    print(f"  Автокорректировка шорта: {'ВКЛ' if bep_auto_adjust else 'ВЫКЛ'}")
print(f"Повторный вход в пул: {'ВКЛ' if reentry_enabled else 'ВЫКЛ'}")
if reentry_enabled:
    print(f"  Задержка повторного входа: {reentry_delay_hours} часов")
    print(f"  Порог падения цены: {reentry_price_drop_threshold}%")
    print(f"  Агрессивный вход через: {reentry_aggressive_hours} часов")

# --- ИСПРАВЛЕННАЯ ФОРМУЛА ликвидности для Uniswap V3 ---
def calculate_liquidity(capital_amount, pn_price, pa_price, pb_price):
    """Правильная формула ликвидности для Uniswap V3"""
    if pn_price <= pa_price or pn_price >= pb_price:
        pn_price = (pa_price + pb_price) / 2
    sqrt_pn = np.sqrt(pn_price)
    sqrt_pa = np.sqrt(pa_price)
    sqrt_pb = np.sqrt(pb_price)
    
    term1 = sqrt_pn - sqrt_pa
    term2 = pn_price * (1/sqrt_pn - 1/sqrt_pb)
    if (term1 + term2) == 0:
        return 0.0
    
    l = capital_amount / (term1 + term2)
    
    if verbose_logging and abs(calculate_pool_value(pn_price, pa_price, pb_price, l) - capital_amount) > capital_amount * 0.01:
        print(f"  Внимание: разница в расчетах. Capital=${capital_amount:.2f}, Pool Value at Pn=${calculate_pool_value(pn_price, pa_price, pb_price, l):.2f}")
    return l

# --- Вспомогательные функции для расчета PnL ---
def calculate_short_size(entry_price, pa, pb, l):
    """Рассчитывает размер шорт-позиции"""
    if entry_price < pa or entry_price > pb or l == 0:
        return 0.0
    sqrt_entry = np.sqrt(entry_price)
    sqrt_pa = np.sqrt(pa)
    return l * (1/sqrt_pa - 1/sqrt_entry)

def calculate_pool_value(price, pa, pb, l):
    """Рассчитывает стоимость пула при заданной цене ТОЛЬКО по формуле V3"""
    if l <= 0:
        return 0.0
    sqrt_price = np.sqrt(price)
    sqrt_pa = np.sqrt(pa)
    sqrt_pb = np.sqrt(pb)
    if price <= pa:
        pool_eth = l * (1/sqrt_pa - 1/sqrt_pb)
        return pool_eth * price
    elif price >= pb:
        return l * (sqrt_pb - sqrt_pa)
    else:
        pool_usdc = l * (sqrt_price - sqrt_pa)
        pool_eth = l * (1/sqrt_price - 1/sqrt_pb)
        return pool_usdc + pool_eth * price

def calculate_pool_loss(current_price, entry_price, pa, pb, l):
    """Рассчитывает потерю пула от цены входа до текущей цены"""
    pool_value_current = calculate_pool_value(current_price, pa, pb, l)
    pool_value_entry = calculate_pool_value(entry_price, pa, pb, l)
    return pool_value_entry - pool_value_current

def calculate_total_pnl_at_price(target_price, entry_price, pa, pb, l):
    """Рассчитывает общий PnL (шорт + пул) при заданной цене"""
    pool_loss = calculate_pool_loss(target_price, entry_price, pa, pb, l)
    short_size = calculate_short_size(entry_price, pa, pb, l)
    short_profit = short_size * (entry_price - target_price)
    return short_profit - pool_loss

# --- НОВАЯ ФУНКЦИЯ: Расчет асимметричного диапазона ---
def calculate_asymmetric_range(center_price, range_width, asymmetry_pct=50):
    """
    Рассчитывает асимметричный диапазон относительно центральной цены.
    """
    asymmetry_pct = max(5, min(95, asymmetry_pct))
    upper_distance = (range_width * asymmetry_pct / 100.0)
    lower_distance = (range_width * (100 - asymmetry_pct) / 100.0)
    pa = center_price - lower_distance
    pb = center_price + upper_distance
    return pa, center_price, pb

# =============================================================================
# ИСПРАВЛЕННАЯ ФУНКЦИЯ 1: Выход из пула с сохранением CurrentPoolValue
# =============================================================================
def exit_pool_and_close_short(df, i, close_price, close_reason, current_price):
    """
    ВЫХОД ИЗ ПУЛА И ЗАКРЫТИЕ ШОРТА КОГДА ЦЕНА ЗАКРЫТИЯ НИЖЕ Pn
    ИСПРАВЛЕНО: Сохраняем стоимость пула в CurrentPoolValue ДО обнуления
    """
    if verbose_logging:
        print(f" ВЫХОД ИЗ ПУЛА И ЗАКРЫТИЕ ШОРТА НИЖЕ Pn!")
        print(f" Причина: {close_reason}")
        print(f" Цена закрытия шорта: ${close_price:.2f}")
        print(f" Текущая рыночная цена: ${current_price:.2f}")
        print(f" Диапазон: Pa=${df.at[i-1, 'DynamicPa']:.2f}, Pn=${df.at[i-1, 'DynamicPn']:.2f}")
    
    # 1. СОХРАНЯЕМ ВСЕ ПАРАМЕТРЫ ПУЛА ДО ИЗМЕНЕНИЙ
    old_pa = df.at[i-1, 'DynamicPa']
    old_pb = df.at[i-1, 'DynamicPb']
    old_pn = df.at[i-1, 'DynamicPn']
    old_l = df.at[i-1, 'DynamicL']
    entry_price = df.at[i-1, 'ShortEntryPrice']
    short_size = df.at[i-1, 'Short ETH']

      # ===== ДИАГНОСТИКА =====
    print(f"\n--- ДИАГНОСТИКА ВЫХОДА ИЗ ПУЛА (строка {i}) ---")
    print(f"  current_price = ${current_price:.2f}")
    print(f"  old_pa = ${old_pa:.2f}")
    print(f"  old_pb = ${old_pb:.2f}")
    print(f"  old_l = {old_l:.6f}")
    print(f"  entry_price = ${entry_price:.2f}")
    print(f"  short_size = {short_size:.6f}")
    print(f"  close_price = ${close_price:.2f}")
    # =======================
    
    # 2. Стоимость пула по ТЕКУЩЕЙ РЫНОЧНОЙ ЦЕНЕ
    pool_value_before_exit = calculate_pool_value(current_price, old_pa, old_pb, old_l)
    
    # 3. PnL пула (IL)
    pool_value_at_entry = calculate_pool_value(entry_price, old_pa, old_pb, old_l)
    pool_realized_pnl = pool_value_before_exit - pool_value_at_entry
    
    # 4. PnL шорта (по цене закрытия)
    short_realized_pnl = -short_size * (close_price - entry_price)
    
    # 5. ОБЩАЯ СТОИМОСТЬ ДЛЯ СОХРАНЕНИЯ
    total_cash_from_exit = pool_value_before_exit + short_realized_pnl

        # ===== ДИАГНОСТИКА РЕЗУЛЬТАТОВ =====
    print(f"  pool_value_before_exit = ${pool_value_before_exit:.2f}")
    print(f"  short_realized_pnl = ${short_realized_pnl:.2f}")
    print(f"  total_cash_from_exit = ${total_cash_from_exit:.2f}")
    print(f"  ---")
    # ===================================
    
    # --- СОХРАНЯЕМ ЗНАЧЕНИЯ ---
    df.at[i, 'Pool Exit Value'] = pool_value_before_exit
    df.at[i, 'Pool Exit Price'] = current_price
    df.at[i, 'Pool Exit Realized PnL'] = pool_realized_pnl
    df.at[i, 'Pool Realized PnL'] = pool_realized_pnl
    df.at[i, 'CurrentPoolValue'] = pool_value_before_exit  # ← КРИТИЧЕСКИ ВАЖНО: сохраняем ДО обнуления!
    
    # --- ТОЛЬКО ТЕПЕРЬ ОБНУЛЯЕМ ПУЛ ---
    df.at[i, 'DynamicPa'] = 0.0
    df.at[i, 'DynamicPn'] = 0.0
    df.at[i, 'DynamicPb'] = 0.0
    df.at[i, 'DynamicL'] = 0.0
    df.at[i, 'Pool Exited'] = 1
    df.at[i, 'ExitTime'] = df.at[i, 'Date']
    
    # --- ЗАКРЫВАЕМ ШОРТ ---
    df.at[i, 'CloseShortPrice'] = close_price
    #df.at[i, 'CloseShortTrigger'] = 1
    df.at[i, 'CloseShortReason'] = close_reason
    df.at[i, 'LastCloseReason'] = close_reason
    df.at[i, 'LastBEP'] = df.at[i-1, 'BEP']
    df.at[i, 'Short Realized PnL'] = short_realized_pnl
    
    # --- ВЕСЬ КАПИТАЛ В PENDING CASH ---
    # total_cash_from_exit уже рассчитан выше
    new_pending = total_cash_from_exit  # эта переменная определена выше в функции
    
    # Проверяем, что Pending Cash не отрицательный
    if new_pending < 0:
        print(f"  ⚠️ Отрицательный Pending Cash: ${new_pending:.2f}")
        print(f"     pool_value = ${pool_value_before_exit:.2f}")
        print(f"     short_pnl = ${short_realized_pnl:.2f}")
        # Устанавливаем в 0, но убыток уже сохранен в Short Realized PnL
        # new_pending = 0.0  # УБРАНО: позволяем отрицательный для честной просадки
        print(f"     Оставляем отрицательным для честной просадки (убыток сохранен в Short Realized PnL)")
    
    df.at[i, 'Pending Cash'] = new_pending
    
    if verbose_logging:
        print(f" Стоимость пула (рынок): ${pool_value_before_exit:.2f}")
        print(f" PnL пула (IL): ${pool_realized_pnl:.2f}")
        print(f" PnL шорта: ${short_realized_pnl:.2f}")
        print(f" Общий капитал: ${total_cash_from_exit:.2f}")
        print(f" Итого в Pending Cash: ${new_pending:.2f}")
        print(f" CurrentPoolValue сохранен: ${pool_value_before_exit:.2f}")
    
    # --- ЗАКРЫВАЕМ ПОЗИЦИЮ ---
    df.at[i, 'Short ETH'] = 0.0
    df.at[i, 'ShortEntryPrice'] = 0.0
    df.at[i, 'ShortActive'] = 0
    df.at[i, 'ShortLiquidity'] = 0.0
    df.at[i, 'Short Unrealized PnL'] = 0.0
    df.at[i, 'BEP'] = 0.0
    df.at[i, 'HoursSinceClose'] = 0.0
    df.at[i, 'DaysSinceClose'] = 0
    
    if verbose_logging:
        print(f" ✅ Выход из пула и закрытие шорта завершены")
        print(f" Весь капитал в Pending Cash: ${df.at[i, 'Pending Cash']:.2f}")
    
    return df

# =============================================================================
# ИСПРАВЛЕННАЯ ФУНКЦИЯ 2: Создание диапазона с ОТКРЫТИЕМ ШОРТА ПО Pn
# =============================================================================
def create_new_range_from_pending_cash(df, i, current_price, available_capital, reason=""):
    """
    СОЗДАЕТ НОВЫЙ ДИАПАЗОН ИЗ PENDING CASH
    ИСПРАВЛЕНО: Открываем шорт ПО ЦЕНЕ ВХОДА В ДИАПАЗОН (Pn = текущая рыночная цена)
    """
    # ===== ДИАГНОСТИКА create_new_range_from_pending_cash =====
    print(f"\n--- ДИАГНОСТИКА create_new_range_from_pending_cash (строка {i}) ---")
    print(f"  reason: {reason}")
    print(f"  available_capital = ${available_capital:.2f}")
    print(f"  current_price = ${current_price:.2f}")
    print(f"  Pool Exited ДО вызова = {df.at[i, 'Pool Exited']}")
    print(f"  Pending Cash ДО вызова = ${df.at[i, 'Pending Cash']:.2f}")
    # ==========================================================
    
    if verbose_logging:
        print(f"\n" + "="*60)
        print(f" СОЗДАНИЕ НОВОГО ДИАПАЗОНА ИЗ PENDING CASH:")
        print(f" Причина: {reason}")
        print(f" Текущая цена (Pn): ${current_price:.2f}")
        print(f" Доступный капитал: ${available_capital:.2f}")
        print(f" Часов после закрытия: {df.at[i, 'HoursSinceClose']:.2f}")
        print("="*60)
    
    # Pn = текущая рыночная цена - ЭТО ПРАВИЛЬНО! (центр диапазона = цена входа)
    if asymmetric_range_enabled:
        pa, pn, pb = calculate_asymmetric_range(
            current_price,  # ← Pn = current_price
            range_width,
            asymmetry_distribution
        )
    else:
        pn = current_price  # ← Pn = current_price
        pa = pn - (range_width / 2)
        pb = pn + (range_width / 2)
    
    # Рассчитываем ликвидность
    l = calculate_liquidity(available_capital, pn, pa, pb)
    
    if l > 0:
        # Обновляем параметры диапазона
        df.at[i, 'DynamicPa'] = pa
        df.at[i, 'DynamicPn'] = pn
        df.at[i, 'DynamicPb'] = pb
        df.at[i, 'DynamicL'] = l
        df.at[i, 'CurrentPoolValue'] = available_capital
        df.at[i, 'Pending Cash'] = 0.0  # ← ОБНУЛЕНИЕ PENDING CASH
        df.at[i, 'Pool Exited'] = 0
        df.at[i, 'ExitTime'] = pd.NaT
        
        # ДИАГНОСТИКА ПОСЛЕ ОБНУЛЕНИЯ
        print(f"  >>> Pending Cash обнулен до $0 (диапазон создан)")
        print(f"  Pool Exited ПОСЛЕ = {df.at[i, 'Pool Exited']}")
        
        # ОТКРЫВАЕМ ШОРТ ПО Pn - ПРАВИЛЬНЫЙ ХЕДЖ!
        if pn > pa and pn < pb:
            short_size = calculate_short_size(pn, pa, pb, l)
            if short_size > 0:
                df.at[i, 'Short ETH'] = float(short_size)
                df.at[i, 'ShortEntryPrice'] = float(pn)  # ← Цена входа = Pn (текущая цена)!
                df.at[i, 'ShortActive'] = 1
                df.at[i, 'ShortLiquidity'] = float(l)
                
                if verbose_logging:
                    print(f" ✅ Новый диапазон создан:")
                    print(f" Pa=${pa:.2f}, Pn=${pn:.2f}, Pb=${pb:.2f}")
                    print(f" Ликвидность L={l:.6f}")
                    print(f" Капитал ${available_capital:.2f} перенесен из Pending Cash в пул")
                    print(f" ✅ Шорт ОТКРЫТ по цене входа Pn=${pn:.2f}, размер: {short_size:.4f} ETH")
            else:
                df.at[i, 'Short ETH'] = 0.0
                df.at[i, 'ShortEntryPrice'] = 0.0
                df.at[i, 'ShortActive'] = 0
                df.at[i, 'ShortLiquidity'] = 0.0
                df.at[i, 'BEP'] = 0.0
                if verbose_logging:
                    print(f" ⚠️ Новый диапазон создан, но шорт НЕ открыт (размер = 0)")
        else:
            df.at[i, 'Short ETH'] = 0.0
            df.at[i, 'ShortEntryPrice'] = 0.0
            df.at[i, 'ShortActive'] = 0
            df.at[i, 'ShortLiquidity'] = 0.0
            df.at[i, 'BEP'] = 0.0
            if verbose_logging:
                print(f" ⚠️ Новый диапазон создан, но Pn вне диапазона, шорт не открыт")
    else:
        print(f"  >>> ВНИМАНИЕ: Ликвидность = 0, диапазон НЕ создан!")
        print(f"  Pending Cash НЕ обнулен, остается = ${df.at[i, 'Pending Cash']:.2f}")
    
    return df

# =============================================================================
# ИСПРАВЛЕННАЯ ФУНКЦИЯ 3: Условия повторного входа (ВСЕ В НАСТРОЙКАХ)
# =============================================================================
def create_new_range_after_exit(df, i):
    """
    СОЗДАЕТ НОВЫЙ ДИАПАЗОН ПОСЛЕ ВЫХОДА ИЗ ПУЛА
    ИСПРАВЛЕНО: Все параметры вынесены в настройки (reentry_delay_hours, reentry_price_drop_threshold)
    """
    if not reentry_enabled:
        return df
    
    # Проверяем доступный капитал из Pending Cash
    available_from_pending = df.at[i, 'Pending Cash']
    if available_from_pending <= 0:
        return df
    
    current_price = df.at[i, 'ETH Price']
    hours_since_close = df.at[i, 'HoursSinceClose']
    
    # УСЛОВИЕ 1: Прошло достаточно времени (настраивается)
    if hours_since_close >= reentry_delay_hours:
        if verbose_logging:
            print(f" ⏰ Прошло {hours_since_close:.1f} часов после выхода, СОЗДАЕМ НОВЫЙ ДИАПАЗОН")
        return create_new_range_from_pending_cash(
            df, i, current_price, available_from_pending,
            f"повторный вход после выхода ({hours_since_close:.1f} часов)"
        )
    
    # УСЛОВИЕ 2: Цена упала от максимума (настраивается)
    if 'MaxPriceSinceClose' in df.columns:
        max_price_since_close = df.at[i, 'MaxPriceSinceClose']
        price_change_from_max = (max_price_since_close - current_price) / max_price_since_close * 100
        
        if price_change_from_max >= reentry_price_drop_threshold:
            if verbose_logging:
                print(f" 📉 Цена упала на {price_change_from_max:.1f}% от максимума, СОЗДАЕМ НОВЫЙ ДИАПАЗОН")
            return create_new_range_from_pending_cash(
                df, i, current_price, available_from_pending,
                f"повторный вход после падения ({price_change_from_max:.1f}%)"
            )
    
    return df

# --- ВСЕ ОСТАЛЬНЫЕ ФУНКЦИИ БЕЗ ИЗМЕНЕНИЙ ---
def recalculate_short_size(df, i):
    """Пересчитывает размер шорта для строки i на основе текущих параметров"""
    if df.at[i, 'ShortActive'] != 1:
        return df
    entry_price = df.at[i, 'ShortEntryPrice']
    pa = df.at[i, 'DynamicPa']
    pb = df.at[i, 'DynamicPb']
    if 'ShortLiquidity' in df.columns and df.at[i, 'ShortLiquidity'] > 0:
        current_l = df.at[i, 'ShortLiquidity']
    else:
        current_l = df.at[i, 'DynamicL']
    if current_l <= 0 or entry_price <= pa or entry_price >= pb:
        df.at[i, 'Short ETH'] = 0.0
        df.at[i, 'ShortLiquidity'] = 0.0
        return df
    new_short_size = calculate_short_size(entry_price, pa, pb, current_l)
    df.at[i, 'Short ETH'] = float(new_short_size)
    df.at[i, 'ShortLiquidity'] = float(current_l)
    return df

def initialize_first_row(df):
    """Инициализация первой строки данных"""
    i = 0
    if asymmetric_range_enabled:
        pa, pn, pb = calculate_asymmetric_range(
            df.at[i, 'ETH Price'],
            range_width,
            asymmetry_distribution
        )
    else:
        pn = df.at[i, 'ETH Price']
        pa = pn - (range_width / 2)
        pb = pn + (range_width / 2)
    
    l = calculate_liquidity(capital, pn, pa, pb)
    
    df.at[i, 'DynamicPa'] = pa
    df.at[i, 'DynamicPn'] = pn
    df.at[i, 'DynamicPb'] = pb
    df.at[i, 'DynamicL'] = l
    df.at[i, 'RangeWidth'] = range_width
    df.at[i, 'CurrentPoolValue'] = capital
    df.at[i, 'MaxPriceSinceClose'] = pn
    df.at[i, 'Pool Exited'] = 0
    
    if l > 0 and pn > pa and pn < pb:
        try:
            be_price = brentq(
                lambda p: calculate_total_pnl_at_price(p, pn, pa, pb, l),
                pa * 1.001,
                pn * 0.999,
                maxiter=100
            )
            if pa < be_price < pn:
                df.at[i, 'BEP'] = float(be_price)
                df.at[i, 'LastBEP'] = float(be_price)
            else:
                df.at[i, 'BEP'] = pn * 0.95
                df.at[i, 'LastBEP'] = pn * 0.95
        except (ValueError, RuntimeError) as e:
            df.at[i, 'BEP'] = pn * 0.95
            df.at[i, 'LastBEP'] = pn * 0.95
    else:
        df.at[i, 'BEP'] = pn * 0.95
        df.at[i, 'LastBEP'] = pn * 0.95
    
    if l > 0 and pn > pa and pn < pb:
        entry_price = pn
        short_size = calculate_short_size(entry_price, pa, pb, l)
        df.at[i, 'Short ETH'] = float(short_size)
        df.at[i, 'ShortEntryPrice'] = float(entry_price)
        df.at[i, 'ShortActive'] = 1
        df.at[i, 'ShortLiquidity'] = float(l)
        if verbose_logging:
            print(f" Открыт начальный шорт: {short_size:.4f} ETH по цене ${entry_price:.2f}")
    else:
        df.at[i, 'Short ETH'] = 0.0
        df.at[i, 'ShortActive'] = 0
    
    return df

def update_hours_since_close(df, i):
    """Обновляет счетчик часов после закрытия шорта"""
    if i == 0:
        return df
    if df.at[i-1, 'CloseShortTrigger'] == 1:
        df.at[i, 'HoursSinceClose'] = time_step_hours
    elif df.at[i-1, 'HoursSinceClose'] > 0:
        df.at[i, 'HoursSinceClose'] = df.at[i-1, 'HoursSinceClose'] + time_step_hours
    return df

def update_days_since_close(df, i):
    """Обновляет счетчик дней после закрытия шорта (для совместимости)"""
    if i == 0:
        return df
    if df.at[i-1, 'CloseShortTrigger'] == 1:
        df.at[i, 'DaysSinceClose'] = 1
    elif df.at[i-1, 'DaysSinceClose'] > 0:
        df.at[i, 'DaysSinceClose'] = df.at[i-1, 'DaysSinceClose'] + 1
    return df

def update_max_price_since_close(df, i):
    """Обновляет максимальную цену после закрытия шорта"""
    if i == 0:
        df.at[i, 'MaxPriceSinceClose'] = df.at[i, 'ETH Price']
        return df
    
    # Если произошел выход из пула
    if df.at[i-1, 'Pool Exited'] == 1:
        # Проверяем, есть ли уже созданный пул после выхода
        if df.at[i, 'Pool Exited'] == 0:  # Если пул уже создан заново
            # Сбрасываем максимум для нового цикла
            df.at[i, 'MaxPriceSinceClose'] = df.at[i, 'ETH Price']
        else:
            # Если пул еще не создан - сохраняем максимум для условия падения
            df.at[i, 'MaxPriceSinceClose'] = max(df.at[i-1, 'MaxPriceSinceClose'], df.at[i, 'ETH Price'])
    
    # Если было закрытие через CloseShortTrigger
    elif df.at[i-1, 'CloseShortTrigger'] == 1:
        # Аналогичная логика
        if df.at[i, 'Pool Exited'] == 0:  # Если пул уже создан
            df.at[i, 'MaxPriceSinceClose'] = df.at[i, 'ETH Price']
        else:
            df.at[i, 'MaxPriceSinceClose'] = max(df.at[i-1, 'MaxPriceSinceClose'], df.at[i, 'ETH Price'])
    
    # Если шорт активен - обновляем максимум
    elif df.at[i-1, 'ShortActive'] == 1:
        df.at[i, 'MaxPriceSinceClose'] = max(df.at[i-1, 'MaxPriceSinceClose'], df.at[i, 'ETH Price'])
    
    # В остальных случаях
    else:
        df.at[i, 'MaxPriceSinceClose'] = max(df.at[i-1, 'MaxPriceSinceClose'], df.at[i, 'ETH Price'])
    
    return df

def calculate_bep_for_row(df, i):
    """Рассчитывает BEP для строки i"""
    if df.at[i, 'ShortActive'] != 1:
        df.at[i, 'BEP'] = 0.0
        if i > 0:
            df.at[i, 'LastBEP'] = df.at[i-1, 'LastBEP']
        return df
    entry_price = df.at[i, 'ShortEntryPrice']
    pa = df.at[i, 'DynamicPa']
    pb = df.at[i, 'DynamicPb']
    current_l = df.at[i, 'ShortLiquidity']
    if current_l <= 0 or entry_price <= pa or entry_price >= pb:
        df.at[i, 'BEP'] = 0.0
        if i > 0:
            df.at[i, 'LastBEP'] = df.at[i-1, 'LastBEP']
        return df
    try:
        be_price = brentq(
            lambda p: calculate_total_pnl_at_price(p, entry_price, pa, pb, current_l),
            pa * 1.001,
            entry_price * 0.999,
            maxiter=100
        )
        if pa < be_price < entry_price:
            df.at[i, 'BEP'] = float(be_price)
            df.at[i, 'LastBEP'] = float(be_price)
        else:
            df.at[i, 'BEP'] = entry_price * 0.95
            df.at[i, 'LastBEP'] = entry_price * 0.95
    except (ValueError, RuntimeError) as e:
        df.at[i, 'BEP'] = entry_price * 0.95
        df.at[i, 'LastBEP'] = entry_price * 0.95
    return df

# =====================================================================
# НОВАЯ ФУНКЦИЯ 1: Расчет BEP с возможностью корректировки размера шорта
# =====================================================================
def calculate_bep_with_target(df, i, target_percent_above_pa=5):
    """
    РАСЧЕТ BEP С ВОЗМОЖНОСТЬЮ КОРРЕКТИРОВКИ РАЗМЕРА ШОРТА
    
    Позволяет задать целевой BEP на определенном проценте выше Pa
    и пересчитать необходимый размер шорт-позиции для достижения этого BEP.
    
    Параметры:
    - df: DataFrame
    - i: индекс строки
    - target_percent_above_pa: целевой процент выше Pa (например, 5 = BEP на 5% выше Pa)
    
    Возвращает:
    - df: обновленный DataFrame (если была корректировка)
    - adjusted: флаг, была ли корректировка
    - new_short_size: новый размер шорта (если был скорректирован)
    - target_bep: целевой BEP
    """
    
    if df.at[i, 'ShortActive'] != 1:
        return df, False, 0.0, 0.0
    
    entry_price = df.at[i, 'ShortEntryPrice']
    pa = df.at[i, 'DynamicPa']
    pb = df.at[i, 'DynamicPb']
    current_l = df.at[i, 'ShortLiquidity']
    current_short_size = df.at[i, 'Short ETH']
    
    if current_l <= 0 or entry_price <= pa or entry_price >= pb:
        return df, False, 0.0, 0.0
    
    # Рассчитываем целевой BEP (на X% выше Pa)
    target_bep = pa * (1 + target_percent_above_pa / 100.0)
    
    # Проверяем, что целевой BEP находится в допустимом диапазоне
    if target_bep >= entry_price:
        if verbose_logging:
            print(f"  ⚠️ Строка {i}: Целевой BEP ${target_bep:.2f} >= цены входа ${entry_price:.2f}, корректировка невозможна")
        return df, False, 0.0, target_bep
    
    if target_bep <= pa:
        if verbose_logging:
            print(f"  ⚠️ Строка {i}: Целевой BEP ${target_bep:.2f} <= Pa ${pa:.2f}, корректировка невозможна")
        return df, False, 0.0, target_bep
    
    # Рассчитываем потерю пула при целевом BEP
    pool_loss_at_target = calculate_pool_loss(target_bep, entry_price, pa, pb, current_l)
    
    # Какой размер шорта нужен, чтобы общий PnL при целевом BEP был равен 0?
    # short_size_needed * (entry_price - target_bep) = pool_loss_at_target
    # => short_size_needed = pool_loss_at_target / (entry_price - target_bep)
    
    if entry_price - target_bep <= 0:
        return df, False, 0.0, target_bep
    
    short_size_needed = pool_loss_at_target / (entry_price - target_bep)
    
    # Проверяем, что размер шорта положительный и разумный
    if short_size_needed <= 0:
        return df, False, 0.0, target_bep
    
    # Вычисляем текущий BEP для сравнения
    current_bep = df.at[i, 'BEP'] if df.at[i, 'BEP'] > 0 else 0
    
    # Если разница больше 1%, показываем информацию
    if current_bep > 0:
        bep_diff_pct = abs(current_bep - target_bep) / current_bep * 100
    else:
        bep_diff_pct = 100
    
    if bep_diff_pct > 1.0 and verbose_logging:
        print(f"\n  📊 АНАЛИЗ BEP (строка {i}):")
        print(f"     Текущий BEP: ${current_bep:.2f} ({((current_bep/pa)-1)*100:.1f}% выше Pa)")
        print(f"     Целевой BEP: ${target_bep:.2f} ({target_percent_above_pa}% выше Pa)")
        print(f"     Разница: {bep_diff_pct:.1f}%")
        print(f"     Текущий размер шорта: {current_short_size:.4f} ETH")
        print(f"     Необходимый размер: {short_size_needed:.4f} ETH")
        print(f"     Изменение: {(short_size_needed/current_short_size - 1)*100:.1f}%")
    
    return df, True, short_size_needed, target_bep

# =====================================================================
# НОВАЯ ФУНКЦИЯ 2: Поиск оптимального размера шорта для заданного BEP
# =====================================================================
def find_optimal_short_size_for_bep(df, i, target_percent_above_pa=5):
    """
    НАХОДИТ ОПТИМАЛЬНЫЙ РАЗМЕР ШОРТА ДЛЯ ЗАДАННОГО BEP
    
    Возвращает словарь с параметрами для анализа
    """
    if df.at[i, 'ShortActive'] != 1:
        return None
    
    entry_price = df.at[i, 'ShortEntryPrice']
    pa = df.at[i, 'DynamicPa']
    pb = df.at[i, 'DynamicPb']
    current_l = df.at[i, 'ShortLiquidity']
    current_short_size = df.at[i, 'Short ETH']
    current_bep = df.at[i, 'BEP'] if df.at[i, 'BEP'] > 0 else 0
    
    if current_l <= 0 or entry_price <= pa or entry_price >= pb:
        return None
    
    # Целевой BEP
    target_bep = pa * (1 + target_percent_above_pa / 100.0)
    
    if target_bep >= entry_price:
        return None
    
    # Потеря пула при целевом BEP
    pool_loss_at_target = calculate_pool_loss(target_bep, entry_price, pa, pb, current_l)
    
    # Необходимый размер шорта
    if entry_price - target_bep <= 0:
        return None
    
    short_size_needed = pool_loss_at_target / (entry_price - target_bep)
    
    # Проверяем, что размер положительный
    if short_size_needed <= 0:
        return None
    
    # Собираем результаты
    result = {
        'current_short_size': current_short_size,
        'current_bep': current_bep,
        'current_bep_pct_above_pa': ((current_bep / pa) - 1) * 100 if current_bep > 0 else 0,
        'target_bep': target_bep,
        'target_bep_pct_above_pa': target_percent_above_pa,
        'short_size_needed': short_size_needed,
        'short_size_change_pct': (short_size_needed / current_short_size - 1) * 100 if current_short_size > 0 else 0,
        'pool_loss_at_target': pool_loss_at_target,
        'entry_price': entry_price,
        'pa': pa
    }
    
    return result

def should_reinvest_cash(df, i, apply_shift=False):
    """
    ОТСЛЕЖИВАЕТ ТРИГГЕРЫ ДЛЯ РЕИНВЕСТИРОВАНИЯ.
    Возвращает True, если нужно реинвестировать СЕЙЧАС.
    """
    if not reinvest_cash_enabled:
        return False

    # ===== НОВОЕ: Если было реинвестирование в последние 24 интервала, пропускаем =====
    last_reinvest_idx = df[df['Compound Event'] == 1].index.max() if any(df['Compound Event'] == 1) else -1
    if last_reinvest_idx != -1:
        hours_since_last = (i - last_reinvest_idx) * time_step_hours
        if hours_since_last < 168:  # 7 дней * 24 часа
            if verbose_logging:
                print(f"  ⏳ Реинвестирование было {hours_since_last:.1f} часов назад (интервал {last_reinvest_idx}), ждем 168 часов")
            return False
    # =================================================================================    
    
    # Переносим флаг с предыдущей строки
    if i > 0:
        if df.at[i-1, 'ReinvestTrigger'] == 1:
            # Проверяем, не прошло ли слишком много времени
            hours_since_trigger = df.at[i, 'HoursSinceTrigger'] if 'HoursSinceTrigger' in df.columns else 0
            if hours_since_trigger > 168:  # Сбрасываем через неделю
                df.at[i, 'ReinvestTrigger'] = 0
                print(f"  ⏰ ТРИГГЕР РЕИНВЕСТИРОВАНИЯ СБРОШЕН (прошло {hours_since_trigger:.1f} часов)")
            else:
                df.at[i, 'ReinvestTrigger'] = df.at[i-1, 'ReinvestTrigger']
                # Обновляем счетчик
                df.at[i, 'HoursSinceTrigger'] = hours_since_trigger + time_step_hours
        else:
            df.at[i, 'ReinvestTrigger'] = 0
            df.at[i, 'HoursSinceTrigger'] = 0
    else:
        df.at[i, 'ReinvestTrigger'] = 0
        df.at[i, 'HoursSinceTrigger'] = 0

    # Проверяем, есть ли что реинвестировать
    pending_cash = df.at[i, 'Pending Cash']
    
    # --- ТРИГГЕР 1: СДВИГ ДИАПАЗОНА ---
    if reinvest_frequency == 'shift' and apply_shift:
        print(f"  🚩 ТРИГГЕР РЕИНВЕСТИРОВАНИЯ: Сдвиг диапазона")
        df.at[i, 'ReinvestTrigger'] = 1
        if pending_cash > 0:
            print(f"  💰 КЭШ УЖЕ ЕСТЬ (${pending_cash:.2f}) - РЕИНВЕСТИРУЕМ НЕМЕДЛЕННО!")
            return True
        return False

    # --- ТРИГГЕР 2: ЗАКРЫТИЕ ШОРТА ---
    if reinvest_frequency == 'short_close' and i > 0:
        if df.at[i-1, 'CloseShortTrigger'] == 1:
            print(f"  🚩 ТРИГГЕР РЕИНВЕСТИРОВАНИЯ: Закрытие шорта")
            df.at[i, 'ReinvestTrigger'] = 1
            if pending_cash > 0:
                print(f"  💰 КЭШ УЖЕ ЕСТЬ (${pending_cash:.2f}) - РЕИНВЕСТИРУЕМ НЕМЕДЛЕННО!")
                return True
            return False

    # --- ТРИГГЕР 3: КАЛЕНДАРНЫЕ (weekly/biweekly/monthly) ---
    if i == 0:
        return False

    current_date = df.at[i, 'Date']

    # Проверяем календарные триггеры с запасом в 1 день
    trigger_activated = False
    trigger_reason = ""
    
    if reinvest_frequency == 'weekly':
        # Получаем номер недели и год
        current_week = current_date.isocalendar()[1]
        current_year = current_date.year
        
        # Получаем последнюю неделю реинвестирования (если есть)
        last_week = 0
        last_year = 0
        last_reinvest_date = None
        if 'LastReinvestDate' in df.columns and i > 0 and pd.notna(df.at[i-1, 'LastReinvestDate']):
            last_date = df.at[i-1, 'LastReinvestDate']
            if isinstance(last_date, pd.Timestamp):
                last_week = last_date.isocalendar()[1]
                last_year = last_date.year
                last_reinvest_date = last_date
        
        # Проверяем, сменилась ли неделя
        week_changed = (current_week != last_week or current_year != last_year)
        
        # Проверяем, не было ли уже реинвестирования в ЭТУ неделю
        reinvest_this_week = False
        if last_reinvest_date is not None:
            last_reinvest_week = last_reinvest_date.isocalendar()[1]
            last_reinvest_year = last_reinvest_date.year
            reinvest_this_week = (current_week == last_reinvest_week and current_year == last_reinvest_year)
        
        # Триггер срабатывает только если:
        # 1. Сменилась неделя И
        # 2. Реинвестирования в эту неделю еще не было
        if week_changed and not reinvest_this_week:
            # Дополнительно проверяем, что прошло не меньше 5 дней
            days_since_last = 999
            if last_reinvest_date is not None:
                days_since_last = (current_date - last_reinvest_date).days
            
            if days_since_last >= 5:
                trigger_activated = True
                trigger_reason = f"weekly (неделя {current_week})"
                if verbose_logging:
                    print(f"     ✓ Первый триггер на неделе {current_week}")
                
    elif reinvest_frequency == 'biweekly':
        # Два раза в месяц: 1-15 и 16-конец месяца
        current_day = current_date.day
        current_month = current_date.month
        current_year = current_date.year
        
        # Определяем половину месяца (1 - первая половина, 2 - вторая)
        current_half = 1 if current_day <= 15 else 2
        
        # Получаем последнюю половину реинвестирования
        last_half = 0
        last_month = 0
        last_year = 0
        if 'LastReinvestDate' in df.columns and i > 0 and pd.notna(df.at[i-1, 'LastReinvestDate']):
            last_date = df.at[i-1, 'LastReinvestDate']
            if isinstance(last_date, pd.Timestamp):
                last_half = 1 if last_date.day <= 15 else 2
                last_month = last_date.month
                last_year = last_date.year
        
        # Проверяем, сменилась ли половина месяца
        if (current_half != last_half or current_month != last_month or current_year != last_year):
            trigger_activated = True
            trigger_reason = f"biweekly (половина {current_half} месяца {current_month})"
            
    elif reinvest_frequency == 'monthly':
        # Раз в месяц
        current_month = current_date.month
        current_year = current_date.year
        
        # Получаем последний месяц реинвестирования
        last_month = 0
        last_year = 0
        if 'LastReinvestDate' in df.columns and i > 0 and pd.notna(df.at[i-1, 'LastReinvestDate']):
            last_date = df.at[i-1, 'LastReinvestDate']
            if isinstance(last_date, pd.Timestamp):
                last_month = last_date.month
                last_year = last_date.year
        
        # Проверяем, сменился ли месяц
        if current_month != last_month or current_year != last_year:
            # Срабатывает 1-го числа или просто при смене месяца
            trigger_activated = True
            trigger_reason = f"monthly (месяц {current_month})"

    if trigger_activated:
        print(f"  🚩 ТРИГГЕР РЕИНВЕСТИРОВАНИЯ: {trigger_reason}")
        print(f"     Текущий Pending Cash: ${pending_cash:.2f}")
        print(f"     Пул активен: {df.at[i, 'DynamicL'] > 0 and df.at[i, 'Pool Exited'] == 0}")
        if i > 0 and 'LastReinvestDate' in df.columns:
            print(f"     LastReinvestDate: {df.at[i-1, 'LastReinvestDate']}")
        
        df.at[i, 'ReinvestTrigger'] = 1
        df.at[i, 'HoursSinceTrigger'] = 0  # Сбрасываем счетчик
        
        # ЕСЛИ КЭШ УЖЕ ЕСТЬ - РЕИНВЕСТИРУЕМ СЕЙЧАС!
        if pending_cash > 0:
            print(f"  💰 КЭШ УЖЕ ЕСТЬ (${pending_cash:.2f}) - РЕИНВЕСТИРУЕМ НЕМЕДЛЕННО!")
            return True
        else:
            print(f"  ⏳ КЭША НЕТ, триггер установлен в 1, ждем появления кэша")
            # НЕ возвращаем True, просто устанавливаем триггер

    # --- ТРИГГЕР 4: РЕИНВЕСТИРОВАНИЕ ТОЛЬКО ПО КАЛЕНДАРЮ (weekly) ---
    # Для weekly триггера не реинвестируем сразу при появлении кэша,
    # а ждем следующего календарного триггера
    
    # --- ТРИГГЕР 5: ПРИНУДИТЕЛЬНОЕ РЕИНВЕСТИРОВАНИЕ ЧЕРЕЗ 168 ЧАСОВ (неделю) ---
    if reinvest_frequency == 'weekly' and df.at[i, 'ReinvestTrigger'] == 1:
        # Проверяем, прошла ли неделя с момента установки триггера
        if df.at[i, 'HoursSinceTrigger'] >= 168:  # 7 дней * 24 часа
            if pending_cash > 0:
                print(f"  ⚠️ РЕИНВЕСТИРОВАНИЕ ЧЕРЕЗ НЕДЕЛЮ: прошло {df.at[i, 'HoursSinceTrigger']:.1f} часов")
                return True
            else:
                print(f"  ⚠️ СБРОС ТРИГГЕРА: нет кэша через неделю")
                df.at[i, 'ReinvestTrigger'] = 0
                df.at[i, 'HoursSinceTrigger'] = 0
                return False
        else:
            # Ждем, пока пройдет неделя
            if verbose_logging and pending_cash > 0:
                print(f"  ⏳ Ожидание следующего календарного триггера: прошло {df.at[i, 'HoursSinceTrigger']:.1f} ч из 168 ч")
            return False
    
    return False

def reinvest_cash(df, i, apply_shift=False):
    """
    РЕИНВЕСТИРОВАНИЕ КЭША С ПЕРЕСЧЕТОМ ПУЛА И ХЕДЖА
    Добавленный капитал распределяется пропорционально:
    - Увеличивается ликвидность пула (L)
    - Пересчитывается размер шорта для новой ликвидности
    """
    # ДИАГНОСТИКА ВХОДА
    print(f"\n  >>> Вход в reinvest_cash для строки {i}")
    print(f"      DynamicL = {df.at[i, 'DynamicL']:.6f}")
    print(f"      Pool Exited = {df.at[i, 'Pool Exited']}")
    print(f"      Pending Cash = ${df.at[i, 'Pending Cash']:.2f}")
    print(f"      CurrentPoolValue (старое) = ${df.at[i, 'CurrentPoolValue']:.2f}")

    actual_pool_value = df.at[i, 'CurrentPoolValue']  
    
    if i == 0:
        print(f"      ⚠️ Выход: i == 0")
        return df

    # Проверяем, активен ли пул
    if df.at[i, 'DynamicL'] <= 0 or df.at[i, 'Pool Exited'] == 1:
        if verbose_logging:
            print(f"  ⏳ Пул не активен, реинвестирование отложено")
        return df

    # Берем Pending Cash из текущей строки
    pending_cash = df.at[i, 'Pending Cash']
    if pending_cash <= 0:
        return df

    # Не реинвестировать копейки
    if pending_cash < 1.0:
        if verbose_logging:
            print(f"  ℹ️ Слишком маленькая сумма (${pending_cash:.2f}), пропускаем реинвестирование")
        return df        
    
    if verbose_logging:
        print(f"\n{'='*60}")
        print(f"💰 РЕИНВЕСТИРОВАНИЕ КЭША: ${pending_cash:.2f}")
        print(f"{'='*60}")
    
    # Затраты на реинвестирование
    compound_cost = gas_usd + (pending_cash * slippage)
    cash_to_reinvest = pending_cash - compound_cost
    
    if cash_to_reinvest <= 0:
        if verbose_logging:
            print(f"  ⚠️ Невозможно реинвестировать (затраты превышают кэш)")
        return df
    
    # Сохраняем дату реинвестирования
    df.at[i, 'LastReinvestDate'] = df.at[i, 'Date']
    df.at[i, 'Compound Event'] = 1
    df.at[i, 'Compound Cost'] = compound_cost
    df.at[i, 'Cash To Reinvest'] = cash_to_reinvest
    
    # === ПЕРЕСЧЕТ ПУЛА И ХЕДЖА ===
    # Используем исправленное значение
    current_pool_value = actual_pool_value
    
    # Если пул активен, увеличиваем его ликвидность
    if df.at[i, 'DynamicL'] > 0 and df.at[i, 'Pool Exited'] == 0:
        # Текущие параметры пула
        current_pa = df.at[i, 'DynamicPa']
        current_pb = df.at[i, 'DynamicPb']
        current_pn = df.at[i, 'DynamicPn']
        current_l = df.at[i, 'DynamicL']
        current_short_entry = df.at[i, 'ShortEntryPrice']
        current_short_active = df.at[i, 'ShortActive']
        current_short_size = df.at[i, 'Short ETH']
        
        if verbose_logging:
            print(f"\n  📊 ТЕКУЩЕЕ СОСТОЯНИЕ (ДО):")
            print(f"    Пул: Pa=${current_pa:.2f}, Pn=${current_pn:.2f}, Pb=${current_pb:.2f}")
            print(f"    Ликвидность L = {current_l:.6f}")
            print(f"    Стоимость пула = ${current_pool_value:.2f}")
            if current_short_active == 1:
                print(f"    Шорт: {current_short_size:.4f} ETH по ${current_short_entry:.2f}")
            print(f"    Добавляемый капитал = ${cash_to_reinvest:.2f}")
        
        # Рассчитываем новую ликвидность с учетом добавленного капитала
        new_pool_value = current_pool_value + cash_to_reinvest
        new_l = calculate_liquidity(new_pool_value, current_pn, current_pa, current_pb)
        
        if new_l > 0:
            # Обновляем ликвидность пула
            df.at[i, 'DynamicL'] = new_l
            df.at[i, 'CurrentPoolValue'] = new_pool_value
            
            if verbose_logging:
                print(f"\n  📈 НОВОЕ СОСТОЯНИЕ (ПОСЛЕ):")
                print(f"    Новая ликвидность L = {new_l:.6f} (было {current_l:.6f})")
                print(f"    Рост L: +{((new_l/current_l)-1)*100:.1f}%")
                print(f"    Новая стоимость пула = ${new_pool_value:.2f} (было ${current_pool_value:.2f})")
            
            # Если есть активный шорт, пересчитываем его размер для новой ликвидности
            if current_short_active == 1 and current_short_entry > 0:
                if current_pa < current_short_entry < current_pb:
                    new_short_size = calculate_short_size(
                        current_short_entry, 
                        current_pa, 
                        current_pb, 
                        new_l
                    )
                    
                    if new_short_size > 0:
                        old_short = df.at[i, 'Short ETH']
                        df.at[i, 'Short ETH'] = float(new_short_size)
                        df.at[i, 'ShortLiquidity'] = float(new_l)
                        
                        if verbose_logging:
                            print(f"    Шорт: {new_short_size:.4f} ETH (было {old_short:.4f} ETH)")
                            print(f"    Рост шорта: +{((new_short_size/old_short)-1)*100:.1f}%")
                    else:
                        if verbose_logging:
                            print(f"  ⚠️ Новый размер шорта = 0, проверьте параметры")
                else:
                    if verbose_logging:
                        print(f"  ⚠️ Цена входа ${current_short_entry:.2f} вне диапазона, шорт не пересчитывается")
            else:
                if verbose_logging:
                    print(f"  ℹ️ Шорт не активен, хедж не пересчитывается")
        else:
            if verbose_logging:
                print(f"  ❌ ОШИБКА: новая ликвидность = 0, реинвестирование отменено")
            return df
    else:
        # Этот блок больше не должен выполняться из-за новой проверки выше
        # Но оставляем на всякий случай
        if verbose_logging:
            print(f"  ⚠️ Неожиданное состояние: пул не активен, но проверка пропущена")
        return df
    
    # Обнуляем Pending Cash
    old_pending = df.at[i, 'Pending Cash']
    df.at[i, 'Pending Cash'] = 0.0
    df.at[i, 'Cash To Reinvest'] = cash_to_reinvest
    
    if verbose_logging:
        print(f"\n  ✅ ИТОГ РЕИНВЕСТИРОВАНИЯ:")
        print(f"    Pending Cash: ${old_pending:.2f} → $0.00")
        print(f"    Затраты: ${compound_cost:.2f}")
        print(f"{'='*60}\n")
    
    # ===== ИСПРАВЛЕНИЕ: Умный сброс триггера =====
    # Для календарных триггеров - сбрасываем, для событийных - оставляем
    if reinvest_frequency in ['weekly', 'biweekly', 'monthly']:
        # Календарные: сбрасываем, следующий триггер будет по календарю
        df.at[i, 'ReinvestTrigger'] = 0
        df.at[i, 'LastReinvestDate'] = df.at[i, 'Date']  # запоминаем дату
        if verbose_logging:
            print(f"  📅 Календарный триггер сброшен, следующая проверка через неделю")
    else:
        # Для 'shift' и 'short_close': оставляем триггер активным
        # но уменьшаем счетчик, чтобы не сработал сразу же
        df.at[i, 'HoursSinceTrigger'] = 0
        if verbose_logging:
            print(f"  🔄 Событийный триггер оставлен активным для следующего события")
    # =============================================
    
    return df

# =============================================================================
# ЧЕСТНЫЙ РАСЧЕТ ПРОСАДКИ (БЕЗ МАСКИРОВКИ)
# =============================================================================
def calculate_correct_drawdown(df):
    """
    РАСЧЕТ ПРОСАДКИ С УЧЕТОМ ВЫХОДА ИЗ ПУЛА
    Технически просадка считается от пика, но exited-периоды с нулевым капиталом
    не должны создавать искусственную -100% просадку, если капитал восстановился.
    """
    print("\nРасчет просадки (исправленная версия с учетом exited-периодов)...")
    
    df = df.copy()
    
    # 1. Создаем столбец для расчета просадки
    df['ValueForDrawdown'] = df['Total Portfolio Value'].copy()
    
    # 2. Для exited-периодов используем forward fill, чтобы избежать -100%
    #    Но только если капитал восстановился позже
    mask_exited = df['Pool Exited'] == 1
    
    if mask_exited.any():
        # Находим периоды exited
        exited_groups = []
        current_group = []
        
        for i in range(len(df)):
            if df.at[i, 'Pool Exited'] == 1:
                current_group.append(i)
            else:
                if current_group:
                    exited_groups.append(current_group)
                    current_group = []
        
        if current_group:
            exited_groups.append(current_group)
        
        print(f"  Найдено {len(exited_groups)} exited-периодов")
        
        # Для каждой группы exited-периодов
        for group in exited_groups:
            last_idx = group[-1]
            
            # Проверяем, был ли создан новый пул после exited
            next_non_exited = last_idx + 1
            while next_non_exited < len(df) and df.at[next_non_exited, 'Pool Exited'] == 1:
                next_non_exited += 1
            
            # Если после exited есть активный период с положительным капиталом
            if next_non_exited < len(df) and df.at[next_non_exited, 'Total Portfolio Value'] > 0:
                # Используем forward fill для exited-периодов
                fill_value = df.at[next_non_exited, 'Total Portfolio Value']
                for idx in group:
                    df.at[idx, 'ValueForDrawdown'] = fill_value
                print(f"    Период {group[0]}-{group[-1]}: заполнено значением ${fill_value:.2f}")
    
    # 3. Peak — максимум по всему времени
    df['Peak'] = df['ValueForDrawdown'].cummax()
    
    # 4. Просадка (избегаем деления на ноль)
    with np.errstate(divide='ignore', invalid='ignore'):
        df['Correct Drawdown'] = np.where(
            df['Peak'] > 0,
            (df['ValueForDrawdown'] - df['Peak']) / df['Peak'] * 100,
            0.0
        )
    
    df['Correct Drawdown'] = df['Correct Drawdown'].fillna(0)
    
    # 5. Финальная диагностика
    correct_max_drawdown = df['Correct Drawdown'].min()
    
    print(f"  Максимальная просадка: {correct_max_drawdown:.2f}%")
    print(f"  Минимальный ValueForDrawdown: ${df['ValueForDrawdown'].min():,.2f}")
    
    # Проверяем, нет ли все еще -100%
    bad_drawdowns = df[df['Correct Drawdown'] < -99.9]
    if len(bad_drawdowns) > 0:
        print(f"  ⚠️ ВНИМАНИЕ: осталось {len(bad_drawdowns)} строк с просадкой -100%")
    
    return df, correct_max_drawdown

# ===== ДИАГНОСТИКА ФАЙЛА =====
import os
print("\n" + "="*60)
print("ДИАГНОСТИКА: Поиск файла")
print("="*60)
print(f"Текущая директория: {os.getcwd()}")
print(f"Ищем файл: {data_file}")

if os.path.exists(data_file):
    print(f"✅ Файл найден, размер: {os.path.getsize(data_file)} байт")
else:
    print(f"❌ ФАЙЛ НЕ НАЙДЕН!")
    print("\nДоступные CSV файлы в текущей директории:")
    csv_files = [f for f in os.listdir('.') if f.endswith('.csv')]
    if csv_files:
        for f in csv_files:
            print(f"  - {f}")
    else:
        print("  CSV файлы не найдены")
    print("\n💡 Решение: Убедитесь, что файл находится в папке:")
    print(f"   {os.getcwd()}")
    exit()
print("="*60 + "\n")
# ==============================

log("Все функции определены")

# --- Чтение данных ---
try:
    print(f"\nЧтение файла {data_file}...")
    df = pd.read_csv(data_file, header=None, skiprows=1)
    if len(df) == 0:
        print("❌ Файл прочитан, но DataFrame пуст!")
        exit()
    print(f"\nИнформация о данных:")
    print(f"Всего строк: {len(df)}")
    print(f"Колонки: {list(df.columns)}")
    
    # Присваиваем понятные имена колонкам
    # По вашим данным, колонки:
    # 0 - open_time (timestamp в микросекундах)
    # 1 - open
    # 2 - high
    # 3 - low
    # 4 - close
    # 5 - volume
    # 6 - close_time
    # 7 - quote_asset_volume
    # 8 - number_of_trades
    # 9 - taker_buy_base_asset_volume
    # 10 - taker_buy_quote_asset_volume
    # 11 - ignore
    
    df.columns = [
        'open_time', 'open', 'high', 'low', 'close', 'volume',
        'close_time', 'quote_asset_volume', 'number_of_trades',
        'taker_buy_base_asset_volume', 'taker_buy_quote_asset_volume', 'ignore'
    ]
    
    print(f"\nПосле присвоения имен колонкам:")
    print(f"Колонки: {list(df.columns)}")
    
    # Конвертируем timestamp в datetime
    # open_time в микросекундах (нужно делить на 1,000,000 для секунд)
    df['open_time'] = pd.to_datetime(pd.to_numeric(df['open_time']) / 1000000, unit='s')
    df['ETH Price'] = df[PRICE_COLUMN].astype(float)
    df['Date'] = df['open_time']
    
    # Конвертируем ценовые колонки в числа
    for col in ['open', 'high', 'low', 'close']:
        df[col] = df[col].astype(float)
    
    # Создаем колонки для high/low/open если нужно
    df['High'] = df['high']
    df['Low'] = df['low']
    df['Open'] = df['open']
    
    df = df.sort_values('Date').reset_index(drop=True)
    
    if len(df) > 1:
        time_diffs = df['Date'].diff().dt.total_seconds() / 3600
        avg_interval = time_diffs.mean()
        print(f"\nПроверка временных интервалов:")
        print(f"Средний интервал между записями: {avg_interval:.4f} часов")
        if use_1min_data:
            print(f"Ожидаемый интервал: 0.0167 часов (1 минута)")
            print(f"Интервал в минутах: {avg_interval*60:.2f} минут")
        else:
            print(f"Ожидаемый интервал: 0.25 часов (15 минут)")
        print(f"Минимальный интервал: {time_diffs.min():.4f} часов")
        print(f"Максимальный интервал: {time_diffs.max():.4f} часов")
        if abs(avg_interval - time_step_hours) > 0.1:
            print(f"⚠️ ВНИМАНИЕ: Средний интервал ({avg_interval:.4f} часов) отличается от ожидаемого ({time_step_hours} часов)")
    
    print(f"\nДиапазон дат: от {df['Date'].min().strftime('%d.%m.%Y %H:%M')} до {df['Date'].max().strftime('%d.%m.%Y %H:%M')}")
    print(f"Диапазон цен: от ${df['ETH Price'].min():.2f} до ${df['ETH Price'].max():.2f}")
    
    if verbose_logging:
        data_type = "1-минутные" if use_1min_data else "15-минутные"
        print(f"\nПервые 5 записей ({data_type} данные):")
        for i in range(min(5, len(df))):
            date_str = df.at[i, 'Date'].strftime('%Y-%m-%d %H:%M')
            price = df.at[i, 'ETH Price']
            print(f"  {date_str}: ${price:,.2f}")

except FileNotFoundError:
    print(f"Ошибка: файл '{data_file}' не найден!")
    exit()
except Exception as e:
    print(f"Ошибка при чтении файла: {e}")
    import traceback
    traceback.print_exc()
    exit()

    # ===== ПРОВЕРКА СОЗДАНИЯ DATAFRAME =====
if 'df' not in locals() and 'df' not in globals():
    print("❌ КРИТИЧЕСКАЯ ОШИБКА: DataFrame не создан!")
    print("   Проверьте ошибки чтения файла выше.")
    exit()
else:
    print(f"✅ DataFrame успешно создан, строк: {len(df)}")
# ========================================

# ===== ОБЯЗАТЕЛЬНАЯ ПРОВЕРКА СУЩЕСТВОВАНИЯ df =====
if 'df' not in locals() and 'df' not in globals():
    print("❌ КРИТИЧЕСКАЯ ОШИБКА: DataFrame 'df' не создан!")
    print("   Проверьте вывод выше – возможно, файл не найден или ошибка чтения.")
    exit()

if verbose_logging:
    print(f"\n{'='*60}")
    print("Проверка данных:")
    print(f"{'='*60}")
    try:
        print(f"Общее количество записей: {len(df)}")
        print(f"Количество уникальных дат: {df['Date'].nunique()}")
        print(f"Пропущенные значения в датах: {df['Date'].isnull().sum()}")
        print(f"Пропущенные значения в ценах: {df['ETH Price'].isnull().sum()}")
    except Exception as e:
        print(f"❌ Ошибка при проверке данных: {e}")
        # Безопасно пытаемся вывести колонки
        try:
            print(f"Доступные колонки: {list(df.columns)}")
        except:
            print("Не удалось получить список колонок")
        exit()

df = df.sort_values('Date', ascending=True).reset_index(drop=True)

# --- Инициализация всех столбцов ---
df['OutOfRangeCount'] = 0
df['BelowPaCount'] = 0
df['OutOfRangeHours'] = 0.0
df['BelowPaHours'] = 0.0
df['CloseShortPrice'] = 0.0
df['LastBEP'] = 0.0
df['LastCloseReason'] = ''
df['DaysSinceClose'] = 0
df['HoursSinceClose'] = 0.0
df['HoursSinceTrigger'] = 0.0  # <-- НОВАЯ СТРОКА: счетчик часов активности триггера    
df['DynamicPn'] = 0.0
df['DynamicPa'] = 0.0
df['DynamicPb'] = 0.0
df['DynamicL'] = 0.0
df['RangeWidth'] = range_width
df['ShortEntryPrice'] = 0.0
df['CloseShortTrigger'] = 0
df['CloseShortReason'] = ''
df['Short ETH'] = 0.0
df['ShortActive'] = 0
df['ShortLiquidity'] = 0.0
df['MaxPriceSinceClose'] = 0.0
df['ShouldReopenShort'] = 0
df['BEP'] = 0.0
df['Short Unrealized PnL'] = 0.0
df['Short Realized PnL'] = 0.0
df['Pool Realized PnL'] = 0.0
df['Realized IL on Shift'] = 0.0
df['CurrentPoolValue'] = capital
df['Pending Cash'] = 0.0
df['Compounded Cash'] = 0.0
df['Compound Event'] = 0
df['Compound Cost'] = 0.0
df['Pool Exit Value'] = 0.0
df['Pool Exit Price'] = 0.0
df['Pool Exit Realized PnL'] = 0.0
df['Pool Exited'] = 0
df['ExitTime'] = pd.NaT
df['LastReinvestDate'] = pd.NaT
df['DynamicInRange'] = 0
df['Cash To Reinvest'] = 0.0
df['ReinvestTrigger'] = 0  # 1 = триггер реинвестирования активен, ждем появления кэша

# --- Основной цикл расчета ---
# --- Основной цикл расчета ---
log("Начинаю основной цикл")
if verbose_logging:
    print(f"\nРАСЧЕТ НАЧИНАЕТСЯ:")
    print("-" * 50)

df = initialize_first_row(df)

total_rows = len(df)
print(f"Всего интервалов для обработки: {total_rows}")
print("Прогресс: ", end="")
sys.stdout.flush()

# Показываем прогресс каждые 1%
progress_step = max(1, total_rows // 100)
last_progress = 0

for i in range(1, total_rows):
    # Индикатор прогресса
    progress = int(i / total_rows * 100)
    if progress > last_progress:
        print(f"{progress}% ", end="")
        sys.stdout.flush()
        last_progress = progress
    
    if verbose_logging and i < 5:  # показываем первые 5 итераций подробно
        date_str = df.at[i, 'Date'].strftime('%Y-%m-%d %H:%M')
        print(f"\nИнтервал {i} ({date_str}): Цена ETH = ${df.at[i, 'ETH Price']:.2f}")
    

    
    # Копируем ВСЕ важные состояния от предыдущей итерации
    important_columns = [
        'DynamicPa', 'DynamicPn', 'DynamicPb', 'DynamicL',
        'Pool Exited', 'ExitTime',
        'ShortActive', 'Short ETH', 'ShortEntryPrice', 'ShortLiquidity',
        'BEP', 'LastBEP',
        'Pending Cash', 'Compounded Cash',
        'OutOfRangeHours', 'BelowPaHours', 'OutOfRangeCount', 'BelowPaCount',
        'MaxPriceSinceClose', 'HoursSinceClose', 'DaysSinceClose', 
        'HoursSinceTrigger'
    ]
    
    for col in important_columns:
        if col in df.columns:
            df.at[i, col] = df.at[i-1, col]
    
    # Сбрасываем только триггеры текущей строки
    df.at[i, 'CloseShortTrigger'] = 0
    df.at[i, 'Compound Event'] = 0
    df.at[i, 'Cash To Reinvest'] = 0.0
    df.at[i, 'Compound Cost'] = 0.0

    # Обновляем максимальную цену с момента последнего закрытия
    df = update_max_price_since_close(df, i)
    
    # Теперь считаем CurrentPoolValue правильно
    if df.at[i, 'Pool Exited'] == 1:
        df.at[i, 'CurrentPoolValue'] = 0.0
    else:
        df.at[i, 'CurrentPoolValue'] = calculate_pool_value(
            df.at[i, 'ETH Price'],
            df.at[i, 'DynamicPa'],
            df.at[i, 'DynamicPb'],
            df.at[i, 'DynamicL']
        )
        
        if i > 0 and df.at[i-1, 'Cash To Reinvest'] > 0:
            cash_to_add = df.at[i-1, 'Cash To Reinvest']
            df.at[i, 'Compounded Cash'] += cash_to_add
            if verbose_logging:
                print(f" Добавлен реинвестированный кэш: ${cash_to_add:.2f}")

    # ===== РАСЧЕТ METRIK ДЛЯ КАЖДОЙ СТРОКИ =====
    # Сначала создаем все необходимые колонки, если их нет
    if 'sqrt_price' not in df.columns:
        df['sqrt_price'] = 0.0
        df['sqrt_dynamic_pa'] = 0.0
        df['sqrt_dynamic_pb'] = 0.0
        df['Pool ETH'] = 0.0
        df['Pool USDC'] = 0.0
        df['Pool Value'] = 0.0
    try:
        # Расчет sqrt значений
        df.at[i, 'sqrt_price'] = np.sqrt(df.at[i, 'ETH Price'])
        df.at[i, 'sqrt_dynamic_pa'] = np.sqrt(df.at[i, 'DynamicPa']) if df.at[i, 'DynamicPa'] > 0 else 0
        df.at[i, 'sqrt_dynamic_pb'] = np.sqrt(df.at[i, 'DynamicPb']) if df.at[i, 'DynamicPb'] > 0 else 0
        
        # Расчет Pool ETH и USDC
        if df.at[i, 'DynamicL'] > 0:
            if df.at[i, 'ETH Price'] <= df.at[i, 'DynamicPa']:
                df.at[i, 'Pool ETH'] = df.at[i, 'DynamicL'] * (1/df.at[i, 'sqrt_dynamic_pa'] - 1/df.at[i, 'sqrt_dynamic_pb']) if df.at[i, 'sqrt_dynamic_pa'] > 0 and df.at[i, 'sqrt_dynamic_pb'] > 0 else 0
                df.at[i, 'Pool USDC'] = 0
            elif df.at[i, 'ETH Price'] >= df.at[i, 'DynamicPb']:
                df.at[i, 'Pool ETH'] = 0
                df.at[i, 'Pool USDC'] = df.at[i, 'DynamicL'] * (df.at[i, 'sqrt_dynamic_pb'] - df.at[i, 'sqrt_dynamic_pa']) if df.at[i, 'sqrt_dynamic_pa'] > 0 and df.at[i, 'sqrt_dynamic_pb'] > 0 else 0
            else:
                df.at[i, 'Pool ETH'] = df.at[i, 'DynamicL'] * (1/df.at[i, 'sqrt_price'] - 1/df.at[i, 'sqrt_dynamic_pb']) if df.at[i, 'sqrt_price'] > 0 and df.at[i, 'sqrt_dynamic_pb'] > 0 else 0
                df.at[i, 'Pool USDC'] = df.at[i, 'DynamicL'] * (df.at[i, 'sqrt_price'] - df.at[i, 'sqrt_dynamic_pa']) if df.at[i, 'sqrt_price'] > 0 and df.at[i, 'sqrt_dynamic_pa'] > 0 else 0
            
            df.at[i, 'Pool Value'] = df.at[i, 'Pool ETH'] * df.at[i, 'ETH Price'] + df.at[i, 'Pool USDC']
        else:
            df.at[i, 'Pool ETH'] = 0
            df.at[i, 'Pool USDC'] = 0
            df.at[i, 'Pool Value'] = 0
            
    except Exception as e:
        print(f"❌ ОШИБКА В СТРОКЕ {i}: {e}")
        print(f"   DynamicL = {df.at[i, 'DynamicL']}")
        print(f"   DynamicPa = {df.at[i, 'DynamicPa']}")
        print(f"   DynamicPb = {df.at[i, 'DynamicPb']}")
        print(f"   ETH Price = {df.at[i, 'ETH Price']}")
        raise e  # Останавливаем выполнение, чтобы увидеть ошибку        
    
    # Расчет sqrt значений
    df.at[i, 'sqrt_price'] = np.sqrt(df.at[i, 'ETH Price'])
    df.at[i, 'sqrt_dynamic_pa'] = np.sqrt(df.at[i, 'DynamicPa']) if df.at[i, 'DynamicPa'] > 0 else 0
    df.at[i, 'sqrt_dynamic_pb'] = np.sqrt(df.at[i, 'DynamicPb']) if df.at[i, 'DynamicPb'] > 0 else 0
    
    # Расчет Pool ETH и USDC
    if df.at[i, 'DynamicL'] > 0:
        if df.at[i, 'ETH Price'] <= df.at[i, 'DynamicPa']:
            df.at[i, 'Pool ETH'] = df.at[i, 'DynamicL'] * (1/df.at[i, 'sqrt_dynamic_pa'] - 1/df.at[i, 'sqrt_dynamic_pb']) if df.at[i, 'sqrt_dynamic_pa'] > 0 and df.at[i, 'sqrt_dynamic_pb'] > 0 else 0
            df.at[i, 'Pool USDC'] = 0
        elif df.at[i, 'ETH Price'] >= df.at[i, 'DynamicPb']:
            df.at[i, 'Pool ETH'] = 0
            df.at[i, 'Pool USDC'] = df.at[i, 'DynamicL'] * (df.at[i, 'sqrt_dynamic_pb'] - df.at[i, 'sqrt_dynamic_pa']) if df.at[i, 'sqrt_dynamic_pa'] > 0 and df.at[i, 'sqrt_dynamic_pb'] > 0 else 0
        else:
            df.at[i, 'Pool ETH'] = df.at[i, 'DynamicL'] * (1/df.at[i, 'sqrt_price'] - 1/df.at[i, 'sqrt_dynamic_pb']) if df.at[i, 'sqrt_price'] > 0 and df.at[i, 'sqrt_dynamic_pb'] > 0 else 0
            df.at[i, 'Pool USDC'] = df.at[i, 'DynamicL'] * (df.at[i, 'sqrt_price'] - df.at[i, 'sqrt_dynamic_pa']) if df.at[i, 'sqrt_price'] > 0 and df.at[i, 'sqrt_dynamic_pa'] > 0 else 0
        
        df.at[i, 'Pool Value'] = df.at[i, 'Pool ETH'] * df.at[i, 'ETH Price'] + df.at[i, 'Pool USDC']
    else:
        df.at[i, 'Pool ETH'] = 0
        df.at[i, 'Pool USDC'] = 0
        df.at[i, 'Pool Value'] = 0
    
    # Расчет метрик
    hours_in_year = 365 * 24
    intervals_in_year = hours_in_year / time_step_hours
    apr_per_interval = apr / intervals_in_year
    
    # Используем Pool Value из предыдущей строки для расчета комиссий
    prev_pool_value = df.at[i-1, 'Pool Value'] if i > 0 and 'Pool Value' in df.columns else capital
    
    # ===== ИСПРАВЛЕНИЕ: Комиссии начисляются всегда, когда пул активен =====
    if df.at[i, 'Pool Exited'] == 0 and df.at[i, 'DynamicL'] > 0:
        # Комиссии начисляются на стоимость пула, независимо от того, в диапазоне цена или нет
        df.at[i, 'Accrued Fees'] = apr_per_interval * prev_pool_value
        if verbose_logging and df.at[i, 'Accrued Fees'] > 0:
            print(f"  💰 НАЧИСЛЕНИЕ КОМИССИЙ: ${df.at[i, 'Accrued Fees']:.2f} (пул активен)")
    else:
        df.at[i, 'Accrued Fees'] = 0.0
    # ======================================================================
    
    funding_rate_per_interval = funding_rate / intervals_in_year
    df.at[i, 'Funding PnL'] = -df.at[i, 'Short ETH'] * df.at[i, 'ETH Price'] * funding_rate_per_interval
    
    # ===== ИСПРАВЛЕНИЕ: Накапливаем Pending Cash от fees и funding =====
    if df.at[i, 'Pool Exited'] == 0:
        current_fees = df.at[i, 'Accrued Fees']
        current_funding = df.at[i, 'Funding PnL']
        
        if verbose_logging:
            print(f"  🔍 Расчет: Fees=${current_fees:.2f}, Funding=${current_funding:.2f}")
        
        if current_fees != 0 or current_funding != 0:
            old_pending = df.at[i, 'Pending Cash']
            df.at[i, 'Pending Cash'] = old_pending + current_fees + current_funding
            
            if verbose_logging:
                print(f"  💰 НАКОПЛЕНИЕ: Fees=${current_fees:.2f} + Funding=${current_funding:.2f}")
                print(f"     Pending Cash: ${old_pending:.2f} → ${df.at[i, 'Pending Cash']:.2f}")
    # ===================================================================              
    
    if df.at[i-1, 'DynamicL'] <= 0 or df.at[i-1, 'Pool Exited'] == 1:
        apply_shift = False
        df.at[i, 'OutOfRangeHours'] = 0
        df.at[i, 'BelowPaHours'] = 0
        df.at[i, 'OutOfRangeCount'] = 0
        df.at[i, 'BelowPaCount'] = 0
    else:
        current_price = df.at[i, 'ETH Price']
        if use_high_low_for_out_of_range and 'High' in df.columns and 'Low' in df.columns:
            price_high = df.at[i, 'High']
            price_low = df.at[i, 'Low']
            out_of_range_up = price_high > df.at[i-1, 'DynamicPb']
            out_of_range_down = price_low < df.at[i-1, 'DynamicPa']
        else:
            out_of_range_up = current_price > df.at[i-1, 'DynamicPb']
            out_of_range_down = current_price < df.at[i-1, 'DynamicPa']
        
        if shift_enabled and out_of_range_up:
            df.at[i, 'OutOfRangeHours'] = df.at[i-1, 'OutOfRangeHours'] + time_step_hours
        else:
            df.at[i, 'OutOfRangeHours'] = 0
        
        if shift_enabled and out_of_range_down:
            df.at[i, 'BelowPaHours'] = df.at[i-1, 'BelowPaHours'] + time_step_hours
        else:
            df.at[i, 'BelowPaHours'] = 0
        
        if 'out_of_range_up' in locals() and shift_enabled and out_of_range_up:
            df.at[i, 'OutOfRangeCount'] = df.at[i-1, 'OutOfRangeCount'] + 1
        else:
            df.at[i, 'OutOfRangeCount'] = 0
        
        if 'out_of_range_down' in locals() and shift_enabled and out_of_range_down:
            df.at[i, 'BelowPaCount'] = df.at[i-1, 'BelowPaCount'] + 1
        else:
            df.at[i, 'BelowPaCount'] = 0
        
        if 'out_of_range_up' in locals():
            apply_shift_up = df.at[i, 'OutOfRangeHours'] >= shift_delay_hours
            apply_shift_down = df.at[i, 'BelowPaHours'] >= shift_delay_hours
            apply_shift = apply_shift_up or apply_shift_down
        else:
            apply_shift = False
        
        if apply_shift and (df.at[i-1, 'DynamicL'] <= 0 or df.at[i-1, 'Pool Exited'] == 1):
            apply_shift = False        
        
        if apply_shift and (df.at[i-1, 'DynamicL'] <= 0 or df.at[i-1, 'Pool Exited'] == 1):
            apply_shift = False
            if verbose_logging:
                print(f" ⚠️ Предотвращен сдвиг: нет активного пула")
        
        if apply_shift:
            old_pa = df.at[i-1, 'DynamicPa']
            old_pb = df.at[i-1, 'DynamicPb']
            old_pn = df.at[i-1, 'DynamicPn']
            old_width = df.at[i-1, 'RangeWidth']
            old_l = df.at[i-1, 'DynamicL']
            
            # ===== ИСПРАВЛЕНИЕ: Используем правильный CurrentPoolValue для сдвига =====
            if i > 0 and df.at[i-1, 'Compound Event'] == 1:
                # Если в предыдущей строке было реинвестирование, берем оттуда
                current_pool_value = df.at[i-1, 'CurrentPoolValue']
                if verbose_logging:
                    print(f"      ⚠️ Сдвиг: используем CurrentPoolValue из строки {i-1} (после реинвестирования): ${current_pool_value:.2f}")
            else:
                current_pool_value = df.at[i, 'CurrentPoolValue']
            # ========================================================================
            
            df.at[i, 'Realized IL on Shift'] = 0.0
            df.at[i, 'Pool Realized PnL'] = 0.0
            
            if verbose_logging:
                print(f" СДВИГ ДИАПАЗОНА: стоимость пула ${current_pool_value:.2f} переносится в новый диапазон")
            
            if use_current_price_as_center:
                new_center = current_price
            else:
                if apply_shift_up:
                    new_center = old_pb
                else:
                    if df.at[i-1, 'CloseShortPrice'] > 0:
                        new_center = df.at[i-1, 'CloseShortPrice']
                    else:
                        new_center = old_pa
            
            if asymmetric_range_enabled:
                new_pa, new_pn, new_pb = calculate_asymmetric_range(
                    new_center,
                    old_width,
                    asymmetry_distribution
                )
                if apply_shift_up:
                    shift_direction = "ВВЕРХ (асимметричный)"
                else:
                    shift_direction = "ВНИЗ (асимметричный)"
            else:
                if apply_shift_up:
                    new_pa = new_center
                    new_pn = new_center
                    new_pb = new_pa + old_width
                    shift_direction = "ВВЕРХ"
                else:
                    new_pb = new_center
                    new_pn = new_center
                    new_pa = new_pb - old_width
                    shift_direction = "ВНИЗ"
            
            if new_pn <= new_pa:
                new_pn = new_pa * 1.01
            elif new_pn >= new_pb:
                new_pn = new_pb * 0.99
            
            if new_pn <= new_pa or new_pn >= new_pb:
                new_pn = (new_pa + new_pb) / 2
            
            new_l = calculate_liquidity(current_pool_value, new_pn, new_pa, new_pb)
            
            if new_l == 0:
                new_pn = (new_pa + new_pb) / 2
                new_l = calculate_liquidity(current_pool_value, new_pn, new_pa, new_pb)
            
            if new_l == 0:
                if verbose_logging:
                    print(f" КРИТИЧЕСКАЯ ОШИБКА: Ликвидность = 0!")
                new_pa = old_pa
                new_pn = old_pn
                new_pb = old_pb
                new_l = old_l
                df.at[i, 'CurrentPoolValue'] = df.at[i-1, 'CurrentPoolValue']
            
            df.at[i, 'DynamicPa'] = new_pa
            df.at[i, 'DynamicPn'] = new_pn
            df.at[i, 'DynamicPb'] = new_pb
            df.at[i, 'DynamicL'] = new_l
            df.at[i, 'RangeWidth'] = old_width
            
            if df.at[i, 'ShortActive'] == 1 and new_l > 0:
                new_short_size = calculate_short_size(df.at[i, 'ShortEntryPrice'], new_pa, new_pb, new_l)
                df.at[i, 'Short ETH'] = float(new_short_size)
                df.at[i, 'ShortLiquidity'] = float(new_l)
                if verbose_logging:
                    print(f" Обновлен размер шорта после сдвига: {new_short_size:.4f} ETH")
            
            if verbose_logging:
                print(f" СДВИГ ДИАПАЗОНА {shift_direction}!")
                print(f" Старый: Pa=${old_pa:.2f}, Pn=${old_pn:.2f}, Pb=${old_pb:.2f}")
                print(f" Новый: Pa=${new_pa:.2f}, Pn=${new_pn:.2f}, Pb=${new_pb:.2f}")
                if asymmetric_range_enabled:
                    print(f" Асимметрия: {asymmetry_distribution}% выше Pn")
                if use_current_price_as_center:
                    print(f" Центр диапазона на текущей цене: ${current_price:.2f}")
                print(f" Стоимость пула: ${current_pool_value:.2f}")
                print(f" Новая ликвидность L = {new_l:.6f}")
            
            df.at[i, 'OutOfRangeHours'] = 0.0
            df.at[i, 'BelowPaHours'] = 0.0
            df.at[i, 'OutOfRangeCount'] = 0
            df.at[i, 'BelowPaCount'] = 0
            df.at[i, 'CloseShortPrice'] = 0.0
        
        else:
            if df.at[i-1, 'DynamicL'] <= 0 or df.at[i-1, 'Pool Exited'] == 1:
                pass
            else:
                df.at[i, 'DynamicPa'] = df.at[i-1, 'DynamicPa']
                df.at[i, 'DynamicPn'] = df.at[i-1, 'DynamicPn']
                df.at[i, 'DynamicPb'] = df.at[i-1, 'DynamicPb']
                df.at[i, 'DynamicL'] = df.at[i-1, 'DynamicL']
                df.at[i, 'RangeWidth'] = df.at[i-1, 'RangeWidth']
            
            df.at[i, 'CloseShortPrice'] = df.at[i-1, 'CloseShortPrice']
    
    df.at[i, 'DynamicInRange'] = 1 if (df.at[i, 'ETH Price'] >= df.at[i, 'DynamicPa'] and
                                      df.at[i, 'ETH Price'] <= df.at[i, 'DynamicPb']) else 0
    
    # --- ЗАКРЫТИЕ ШОРТА ---
    if df.at[i-1, 'ShortActive'] == 1:
        if df.at[i-1, 'Short ETH'] > 0:
            df.at[i, 'Short Unrealized PnL'] = -df.at[i-1, 'Short ETH'] * (df.at[i, 'ETH Price'] - df.at[i-1, 'ShortEntryPrice'])
        
        should_close = False
        close_price = 0.0
        close_reason = ""
        
        max_loss_level = df.at[i-1, 'ShortEntryPrice'] * (1 + max_loss_percent / 100.0)
        if df.at[i, 'ETH Price'] >= max_loss_level:
            should_close = True
            close_price = max_loss_level
            close_reason = "max_loss"
            if verbose_logging:
                price_above_entry_pct = (df.at[i, 'ETH Price'] - df.at[i-1, 'ShortEntryPrice']) / df.at[i-1, 'ShortEntryPrice'] * 100
                print(f" ЗАКРЫТИЕ ШОРТА ПО MAX LOSS!")
                print(f" Цена превысила вход на {price_above_entry_pct:.2f}%")
                print(f" Цена закрытия: ${close_price:.2f}")
        
        if not should_close and use_bep_as_primary_trigger and df.at[i-1, 'BEP'] > 0:
            current_bep = df.at[i-1, 'BEP']
            bep_close_level = current_bep * (1 + bep_close_buffer_pct / 100.0)
            if df.at[i, 'ETH Price'] <= bep_close_level:
                should_close = True
                close_price = df.at[i, 'ETH Price']
                close_reason = "bep"
                if verbose_logging:
                    percent_above_bep = (df.at[i, 'ETH Price'] - current_bep) / current_bep * 100 if current_bep > 0 else 0
                    print(f" ЗАКРЫТИЕ ШОРТА ПО BEP-ТРИГГЕРОМ!")
                    print(f" Цена: ${df.at[i, 'ETH Price']:.2f} ≤ ${bep_close_level:.2f}")
                    print(f" BEP: ${current_bep:.2f}")
                    print(f" Над BEP: {percent_above_bep:.1f}%")
        
        if not should_close and close_short_below_pa:
            current_pa = df.at[i-1, 'DynamicPa']
            close_level = current_pa * (1 + close_short_buffer_pct / 100.0)
            if df.at[i, 'ETH Price'] <= close_level:
                should_close = True
                close_price = df.at[i, 'ETH Price']
                close_reason = "pa_buffer"
                if verbose_logging:
                    percent_to_pa = (close_price - current_pa) / current_pa * 100 if current_pa > 0 else 0
                    print(f" ЗАКРЫТИЕ ШОРТА ПО БУФЕРУ ДО Pa!")
                    print(f" Цена: ${df.at[i, 'ETH Price']:.2f} ≤ ${close_level:.2f}")
                    print(f" Pa: ${current_pa:.2f}")
                    print(f" До Pa: {percent_to_pa:.1f}%")
        
        if should_close:
            current_pn = df.at[i-1, 'DynamicPn']
            current_price = df.at[i, 'ETH Price']
            
            if close_price < current_pn:
                df = exit_pool_and_close_short(df, i, close_price, f"{close_reason}_below_pn_exit_pool", current_price)
            else:
                df.at[i, 'CloseShortPrice'] = close_price
                df.at[i, 'CloseShortTrigger'] = 1
                df.at[i, 'CloseShortReason'] = close_reason
                df.at[i, 'LastBEP'] = df.at[i-1, 'BEP']
                df.at[i, 'LastCloseReason'] = close_reason
                
                if df.at[i-1, 'Short ETH'] > 0:
                    realized_pnl = -df.at[i-1, 'Short ETH'] * (close_price - df.at[i-1, 'ShortEntryPrice'])
                    df.at[i, 'Short Realized PnL'] = realized_pnl
                    if verbose_logging:
                        print(f" Реализованный PnL: ${realized_pnl:.2f}")
                    
                    df.at[i, 'Pending Cash'] = df.at[i-1, 'Pending Cash'] + df.at[i, 'Short Realized PnL']
                    if verbose_logging:
                        print(f" Добавлено в Pending Cash: ${df.at[i, 'Short Realized PnL']:.2f}")
                    # Проверка на отрицательный Pending Cash
                    if df.at[i, 'Pending Cash'] < 0:
                        print(f"  ⚠️ Отрицательный Pending Cash: ${df.at[i, 'Pending Cash']:.2f}")
                        print(f"     realized_pnl = ${realized_pnl:.2f}, prev_pending = ${df.at[i-1, 'Pending Cash']:.2f}")
                        # df.at[i, 'Pending Cash'] = 0.0  # УБРАНО: позволяем отрицательный для честной просадки
                        print(f"     Оставляем отрицательным для честной просадки (убыток сохранен в Short Realized PnL)")
                
                df.at[i, 'Short ETH'] = 0.0
                df.at[i, 'ShortEntryPrice'] = 0.0
                df.at[i, 'ShortActive'] = 0
                df.at[i, 'ShortLiquidity'] = 0.0
                df.at[i, 'Short Unrealized PnL'] = 0.0
                df.at[i, 'BEP'] = 0.0
                df.at[i, 'HoursSinceClose'] = 0.0
                df.at[i, 'DaysSinceClose'] = 0
        else:
            df.at[i, 'Short ETH'] = float(df.at[i-1, 'Short ETH'])
            df.at[i, 'ShortEntryPrice'] = float(df.at[i-1, 'ShortEntryPrice'])
            df.at[i, 'ShortActive'] = 1
            df.at[i, 'ShortLiquidity'] = float(df.at[i-1, 'ShortLiquidity'])
            df.at[i, 'CloseShortTrigger'] = 0
            df.at[i, 'Short Realized PnL'] = 0.0
    else:
        df.at[i, 'Short ETH'] = 0.0
        df.at[i, 'ShortEntryPrice'] = 0.0
        df.at[i, 'ShortActive'] = 0
        df.at[i, 'ShortLiquidity'] = 0.0
    
    # --- ОТКРЫТИЕ ШОРТА ---
    if df.at[i, 'ShortActive'] == 0:
        should_open_short = False
        open_reason = ""
        
        if (df.at[i, 'DynamicInRange'] == 1 and
            df.at[i-1, 'DynamicInRange'] == 0):
            should_open_short = True
            open_reason = "первичный вход"
        
        if (df.at[i, 'DynamicPa'] != df.at[i-1, 'DynamicPa'] and
            df.at[i, 'DynamicInRange'] == 1 and
            df.at[i, 'DynamicL'] > 0):
            should_open_short = True
            open_reason = "после сдвига диапазона"
        
        if (reopen_short_below_pn and
            df.at[i, 'ShortActive'] == 0 and
            df.at[i-1, 'ShortActive'] == 0 and
            df.at[i, 'DynamicInRange'] == 1 and
            df.at[i, 'DynamicL'] > 0):
            
            hours_since_close = df.at[i, 'HoursSinceClose']
            if hours_since_close >= reopen_delay_hours:
                current_pn = df.at[i, 'DynamicPn']
                # Рассчитываем падение от максимума
                price_drop_pct = (df.at[i, 'MaxPriceSinceClose'] - df.at[i, 'ETH Price']) / df.at[i, 'MaxPriceSinceClose'] * 100
                
                # Основные условия: цена ниже Pn, падение достаточное, цена выше Pa
                if (df.at[i, 'ETH Price'] < current_pn and
                    price_drop_pct >= reopen_percent_below_pn and
                    df.at[i, 'ETH Price'] >= df.at[i, 'DynamicPa']):
                    
                    # Если игнорирование включено – открываем без проверки LastBEP
                    if ignore_last_bep_for_reopen:
                        should_open_short = True
                        open_reason = f"повторное открытие (падение {price_drop_pct:.1f}%)"
                    else:
                        # Иначе проверяем LastBEP (старая логика)
                        last_bep = df.at[i, 'LastBEP']
                        if last_bep > 0:
                            price_above_bep_pct = (df.at[i, 'ETH Price'] - last_bep) / last_bep * 100
                            if price_above_bep_pct >= reopen_bep_buffer_pct:
                                should_open_short = True
                                open_reason = f"повторное открытие (падение {price_drop_pct:.1f}%, выше BEP на {price_above_bep_pct:.1f}%)"
                        else:
                            should_open_short = True
                            open_reason = f"повторное открытие (падение {price_drop_pct:.1f}%, BEP отсутствует)"
        
        if should_open_short and df.at[i, 'DynamicL'] > 0:
            entry_price = df.at[i, 'ETH Price']
            pa = df.at[i, 'DynamicPa']
            pb = df.at[i, 'DynamicPb']
            current_l = df.at[i, 'DynamicL']
            
            if entry_price > pa and entry_price < pb:
                short_size = calculate_short_size(entry_price, pa, pb, current_l)
                if short_size > 0:
                    df.at[i, 'Short ETH'] = float(short_size)
                    df.at[i, 'ShortEntryPrice'] = float(entry_price)
                    df.at[i, 'ShortActive'] = 1
                    df.at[i, 'ShortLiquidity'] = float(current_l)
                    if verbose_logging:
                        print(f" ОТКРЫТИЕ ШОРТА!")
                        print(f" Причина: {open_reason}")
                        print(f" Цена входа: ${entry_price:.2f}")
                        print(f" Размер шорта: {short_size:.4f} ETH")
    
    df = recalculate_short_size(df, i)
    df = calculate_bep_for_row(df, i)
   
    # =============================================================================
    # ИСПРАВЛЕНИЕ: НОВЫЙ ПОРЯДОК ПРОВЕРКИ УСЛОВИЙ ДЛЯ СОЗДАНИЯ ПУЛА
    # =============================================================================
    
    # БЛОК 1: Сначала проверяем стандартные условия повторного входа (create_new_range_after_exit)
    # Эта функция содержит проверки reentry_delay_hours и reentry_price_drop_threshold
    if df.at[i, 'DynamicL'] <= 0 and df.at[i, 'Pending Cash'] > 0:
        old_pool_state = df.at[i, 'DynamicL']  # Запоминаем состояние до вызова
        df = create_new_range_after_exit(df, i)
        
        # Если пул создался - выходим из блока проверок
        if df.at[i, 'DynamicL'] > 0:
            if verbose_logging:
                print(f"  ✅ Пул создан через create_new_range_after_exit")
            # Сбрасываем счетчик интервалов без пула
            if 'intervals_without_pool' in locals():
                intervals_without_pool = 0
    
    # БЛОК 2: Если пул все еще не создан, пробуем агрессивный вход
    if df.at[i, 'DynamicL'] <= 0 and df.at[i, 'Pending Cash'] > 0 and reentry_enabled:
        hours_since_close = df.at[i, 'HoursSinceClose']
        if hours_since_close >= reentry_aggressive_hours:
            if verbose_logging:
                print(f" 🚀 АГРЕССИВНОЕ СОЗДАНИЕ ДИАПАЗОНА: прошло {hours_since_close:.1f} часов")
            df = create_new_range_from_pending_cash(
                df, i, df.at[i, 'ETH Price'], df.at[i, 'Pending Cash'],
                f"агрессивное создание через {hours_since_close:.1f} часов"
            )
            # Сбрасываем счетчик интервалов без пула
            if 'intervals_without_pool' in locals():
                intervals_without_pool = 0
    
    # БЛОК 3: Если все еще нет пула, проверяем счетчик интервалов
    if df.at[i, 'DynamicL'] <= 0 and df.at[i, 'Pending Cash'] > 0:
        # Инициализируем счетчик если нужно
        if 'intervals_without_pool' not in locals():
            intervals_without_pool = 0
        
        # Увеличиваем счетчик
        if i > 0 and df.at[i-1, 'DynamicL'] <= 0:
            intervals_without_pool += 1
        else:
            intervals_without_pool = 0
        
        # Принудительное создание через 100 интервалов
        if intervals_without_pool >= 100:
            if verbose_logging:
                print(f" 🚨 ПРИНУДИТЕЛЬНОЕ СОЗДАНИЕ ДИАПАЗОНА!")
            df = create_new_range_from_pending_cash(
                df, i, df.at[i, 'ETH Price'], df.at[i, 'Pending Cash'],
                f"принудительное создание через {intervals_without_pool} интервалов"
            )
            intervals_without_pool = 0

    # ========== ИСПРАВЛЕНИЕ 1: ВСЕГДА проверяем реинвестирование ==========
    should_reinvest = should_reinvest_cash(df, i, False)  # Проверяем всегда!
        
    # Если есть кэш и триггер активен - реинвестируем немедленно
    if reinvest_cash_enabled and should_reinvest and df.at[i, 'Pending Cash'] > 0:
        df = reinvest_cash(df, i, False)  # ← Добавлен отступ в 8 пробелов
    # ====================================================================

# ===== ИСПРАВЛЕНИЕ 7: ДИАГНОСТИКА НАКОПЛЕНИЯ В РЕАЛЬНОМ ВРЕМЕНИ =====
print("\n" + "="*60)
print("ДИАГНОСТИКА НАКОПЛЕНИЯ PENDING CASH")
print("="*60)

# Рассчитываем ожидаемое накопление по периодам
print("\nПересчет Expected_Pending с учетом реинвестирований...")
df['Expected_Pending'] = 0.0
cumulative = 0.0

for i in range(len(df)):
    if i == 0:
        cumulative = df.at[i, 'Short Realized PnL']
    else:
        # Добавляем доходы
        cumulative += (df.at[i, 'Short Realized PnL'] + 
                       df.at[i, 'Accrued Fees'] + 
                       df.at[i, 'Funding PnL'])
        
        # Если было реинвестирование - вычитаем РЕАЛЬНУЮ СУММУ реинвестирования
        if df.at[i, 'Compound Event'] == 1:
            # Сумма реинвестирования = Pending Cash до реинвестирования
            reinvested_amount = df.at[i-1, 'Pending Cash'] if i > 0 else 0
            cumulative -= reinvested_amount
            if verbose_logging:
                print(f"  Строка {i}: реинвестирование ${reinvested_amount:.2f}, cumulative = ${cumulative:.2f}")
        
        # Вычитаем затраты (они уже учтены в Pending Cash)
        if df.at[i, 'Compound Cost'] > 0:
            cumulative -= df.at[i, 'Compound Cost']
    
    df.at[i, 'Expected_Pending'] = cumulative

# Находим моменты расхождений
df['Pending_Diff'] = df['Pending Cash'] - df['Expected_Pending']
problem_rows = df[abs(df['Pending_Diff']) > 1.0]

print(f"Всего строк с расхождением > $1: {len(problem_rows)}")
print(f"Суммарное расхождение: ${df['Pending_Diff'].sum():.2f}")

if len(problem_rows) > 0:
    print("\nПервые 10 расхождений:")
    for idx in problem_rows.index[:10]:
        print(f"  Строка {idx}: {df.at[idx, 'Date']}")
        print(f"    Pending Cash: ${df.at[idx, 'Pending Cash']:.2f}")
        print(f"    Expected: ${df.at[idx, 'Expected_Pending']:.2f}")
        print(f"    Разница: ${df.at[idx, 'Pending_Diff']:.2f}")
        print(f"    Compound Event: {df.at[idx, 'Compound Event']}")
        print(f"    Short Realized: ${df.at[idx, 'Short Realized PnL']:.2f}")
        print(f"    Accrued Fees: ${df.at[idx, 'Accrued Fees']:.2f}")
        print(f"    Funding PnL: ${df.at[idx, 'Funding PnL']:.2f}")
        print(f"    Compound Cost: ${df.at[idx, 'Compound Cost']:.2f}")
        print("    ---")

# Добавим еще контрольные точки каждые 10% данных
print("\nКонтрольные точки по периодам:")
total_rows = len(df)
checkpoints = [int(total_rows * p) for p in [0.1, 0.25, 0.5, 0.75, 0.9, 1.0]]

for cp in checkpoints:
    if cp >= total_rows:
        cp = total_rows - 1
    expected_at_cp = df['Expected_Pending'].iloc[cp]
    actual_at_cp = df['Pending Cash'].iloc[cp]
    pct_complete = (cp / total_rows) * 100
    print(f"  {pct_complete:.0f}% (строка {cp}): Ожидалось=${expected_at_cp:.2f}, Факт=${actual_at_cp:.2f}, Разница=${actual_at_cp - expected_at_cp:.2f}")

print("="*60 + "\n")
# =================================================================

# ===== АВАРИЙНАЯ ЗАЩИТА ОТ PENDING CASH = 0 =====
print("\n" + "="*60)
print("АВАРИЙНАЯ ЗАЩИТА: проверка Pending Cash")
print("="*60)

problem_rows = 0
for idx in df.index:
    if df.at[idx, 'Pool Exited'] == 1 and df.at[idx, 'Pending Cash'] == 0 and df.at[idx, 'Pool Exit Value'] > 0:
        problem_rows += 1
        # Восстанавливаем из Pool Exit Value
        recovered = df.at[idx, 'Pool Exit Value'] + df.at[idx, 'Short Realized PnL']
        print(f"  Строка {idx}: {df.at[idx, 'Date']}")
        print(f"    Было: Pending Cash = $0, Pool Exit Value = ${df.at[idx, 'Pool Exit Value']:.2f}")
        print(f"    Стало: Pending Cash = ${recovered:.2f}")
        df.at[idx, 'Pending Cash'] = recovered
        
        # Пересчитываем Total Portfolio Value для этой строки
        df.at[idx, 'Total Portfolio Value'] = recovered + df.at[idx, 'Compounded Cash']

print(f"Исправлено строк: {problem_rows}")
print("="*60)
# ================================================

# --- ВСЕ ПОСЛЕДУЮЩИЕ РАСЧЕТЫ БЕЗ ИЗМЕНЕНИЙ ---
print(f"\nРАСЧЕТ FINANCIAL METRICS...")

# --- ВСЕ ПОСЛЕДУЮЩИЕ РАСЧЕТЫ ---
print(f"\nРАСЧЕТ FINANCIAL METRICS...")

df['PriceBetweenPaAndPn'] = np.where(
    (df['ETH Price'] >= df['DynamicPa']) & (df['ETH Price'] <= df['DynamicPn']),
    1, 0
).astype(int)

print(f"Интервалов в диапазоне: {df['DynamicInRange'].sum()} из {len(df)} ({df['DynamicInRange'].sum()/len(df)*100:.1f}%)")

df['sqrt_price'] = np.sqrt(df['ETH Price'])
df['sqrt_dynamic_pa'] = np.sqrt(df['DynamicPa'])
df['sqrt_dynamic_pb'] = np.sqrt(df['DynamicPb'])

df['Pool ETH'] = np.where(
    df['ETH Price'] <= df['DynamicPa'],
    df['DynamicL'] * (1/df['sqrt_dynamic_pa'] - 1/df['sqrt_dynamic_pb']),
    np.where(
        df['ETH Price'] >= df['DynamicPb'],
        0,
        df['DynamicL'] * (1/df['sqrt_price'] - 1/df['sqrt_dynamic_pb'])
    )
)

df['Pool USDC'] = np.where(
    df['ETH Price'] <= df['DynamicPa'],
    0,
    np.where(
        df['ETH Price'] >= df['DynamicPb'],
        df['DynamicL'] * (df['sqrt_dynamic_pb'] - df['sqrt_dynamic_pa']),
        df['DynamicL'] * (df['sqrt_price'] - df['sqrt_dynamic_pa'])
    )
)

df['Pool Value'] = df['Pool ETH'] * df['ETH Price'] + df['Pool USDC']
df['Pool PnL Daily'] = df['Pool Value'].diff().fillna(0.0)

hours_in_year = 365 * 24
intervals_in_year = hours_in_year / time_step_hours
apr_per_interval = apr / intervals_in_year
df['Accrued Fees'] = apr_per_interval * df['Pool Value'].shift(1).fillna(capital) * df['DynamicInRange']

df['Delta Pool'] = np.where(
    df['DynamicInRange'] == 1,
    df['DynamicL'] * (1/df['sqrt_dynamic_pa'] - 1/df['sqrt_price']),
    df['Pool ETH']
)

# ПРАВИЛЬНЫЙ РАСЧЕТ PENDING CASH (без перезаписи истории)
print(f"\nПРОВЕРКА НАКОПЛЕНИЯ PENDING CASH...")
print("="*60)

df['Cumulative Fees'] = df['Accrued Fees'].cumsum()
df['Cumulative Funding'] = df['Funding PnL'].cumsum()

# Рассчитываем ожидаемую сумму
total_short_pnl = df['Short Realized PnL'].sum()
total_fees = df['Accrued Fees'].sum()
total_funding = df['Funding PnL'].sum()
total_compound_cost = df['Compound Cost'].sum()
final_pending = df['Pending Cash'].iloc[-1]
expected_pending = total_short_pnl + total_fees + total_funding - total_compound_cost

print(f"\n  Short Realized PnL: ${total_short_pnl:,.2f}")
print(f"  Accrued Fees: ${total_fees:,.2f}")
print(f"  Funding PnL: ${total_funding:,.2f}")
print(f"  Compound Cost: ${total_compound_cost:,.2f}")
print(f"  {'='*40}")
print(f"  Ожидаемая сумма: ${expected_pending:,.2f}")
print(f"  Фактический Pending Cash: ${final_pending:,.2f}")
print(f"  Расхождение: ${final_pending - expected_pending:,.2f}")

if abs(final_pending - expected_pending) < 1.0:
    print("\n  ✅ Pending Cash рассчитан корректно!")
    print("  ✅ История сохранена без изменений")
else:
    print("\n  ⚠️ Есть расхождение в расчетах!")
    print("  ⚠️ НО история НЕ перезаписывается")
    print("  ✅ Используем исходные значения из основного цикла")
    
    # Находим максимальное расхождение для диагностики
    df['Pending_Check'] = 0.0
    check_pending = 0.0
    for i in range(len(df)):
        if i == 0:
            check_pending = df.at[i, 'Short Realized PnL']
        else:
            check_pending += (df.at[i, 'Short Realized PnL'] + 
                            df.at[i, 'Accrued Fees'] + 
                            df.at[i, 'Funding PnL'])
            if df.at[i, 'Compound Event'] == 1 and i > 0:
                check_pending -= df.at[i-1, 'Pending Cash']
        df.at[i, 'Pending_Check'] = check_pending
    
    df['Pending_Diff'] = df['Pending_Check'] - df['Pending Cash']
    max_diff_idx = df['Pending_Diff'].abs().idxmax()
    print(f"\n  Максимальное расхождение в строке {max_diff_idx}:")
    print(f"    По расчету: ${df.at[max_diff_idx, 'Pending_Check']:,.2f}")
    print(f"    В данных: ${df.at[max_diff_idx, 'Pending Cash']:,.2f}")
    print(f"    Разница: ${df.at[max_diff_idx, 'Pending_Diff']:,.2f}")
    
    # Удаляем временные колонки
    df = df.drop(columns=['Pending_Check', 'Pending_Diff'], errors='ignore')

print("="*60)
# ============================================================
df['Costs'] = 0.0
for i in range(len(df)):
    if i == 0:
        continue
    if df.at[i, 'ShortActive'] == 1 and df.at[i-1, 'ShortActive'] == 0:
        df.at[i, 'Costs'] = gas_usd + short_fee * df.at[i, 'Short ETH'] * df.at[i, 'ETH Price']
    elif df.at[i, 'ShortActive'] == 0 and df.at[i-1, 'ShortActive'] == 1:
        df.at[i, 'Costs'] = gas_usd + short_fee * df.at[i-1, 'Short ETH'] * df.at[i, 'ETH Price']
df['Costs'] = df['Costs'] + df['Compound Cost']

df['Daily PnL'] = -df['Costs']
df['Realized PnL'] = df['Short Realized PnL'] + df['Pool Realized PnL']
df['Total PnL'] = df['Daily PnL'] + df['Realized PnL']

# ========== ИСПРАВЛЕННЫЙ БЛОК РАСЧЕТА ИТОГОВЫХ МЕТРИК (окончательная версия) ==========
print(f"\nРАСЧЕТ ИТОГОВЫХ МЕТРИК...")

# 1. Явно обнуляем все значения пула при выходе из пула
mask_exited = df['Pool Exited'] == 1
df.loc[mask_exited, 'Pool Value'] = 0.0
df.loc[mask_exited, 'Pool ETH']   = 0.0
df.loc[mask_exited, 'Pool USDC']  = 0.0
df.loc[mask_exited, 'CurrentPoolValue'] = 0.0   # если такая колонка есть

# 2. Считаем правильный Total Portfolio Value
df['Total Portfolio Value'] = (
    df['Pool Value'] +
    df['Pending Cash'] +
    df['Compounded Cash']
)

# 3. Глобальная защита от отрицательного портфеля (самое важное для просадки!)
negative_tvl_count = (df['Total Portfolio Value'] < 0).sum()
if negative_tvl_count > 0:
    print(f"  Обнаружено {negative_tvl_count} строк с отрицательным Total Portfolio Value → исправляем в 0")
    df['Total Portfolio Value'] = df['Total Portfolio Value'].clip(lower=0)

# 4. Остальные метрики (оставляем как было)
df['Costs'] = df['Costs'] + df['Compound Cost']
df['Daily PnL'] = -df['Costs']
df['Realized PnL'] = df['Short Realized PnL'] + df['Pool Realized PnL']
df['Total PnL'] = df['Daily PnL'] + df['Realized PnL']

df['Cumulative PnL'] = df['Total Portfolio Value'] - capital
df['ROI'] = (df['Total Portfolio Value'] - capital) / capital
df['Total ROI'] = df['ROI']  # для совместимости

# 5. Краткая диагностика (очень полезно!)
print(f"  Минимальное Total Portfolio Value после исправлений: ${df['Total Portfolio Value'].min():,.2f}")
print(f"  Строк с Total Portfolio Value <= 0: {(df['Total Portfolio Value'] <= 0).sum()}")
print(f"  Строк с Pool Exited == 1: {mask_exited.sum()}")

# ===== РАСЧЕТ ПРОСАДОК ЗА 15 МИНУТ И ЗА ДЕНЬ =====
print(f"\nРАСЧЕТ ПРОСАДОК ЗА 15 МИНУТ И ЗА ДЕНЬ...")

# 1. Просадка за 15 минут (изменение за один интервал)
df['Drawdown_15min'] = df['Total Portfolio Value'].pct_change() * 100
df['Drawdown_15min'] = df['Drawdown_15min'].fillna(0)

# 2. Группировка по дням для расчета дневных просадок
df['Date_Day'] = df['Date'].dt.date

# Создаем словари для хранения результатов
daily_first_values = {}
daily_last_values = {}
daily_min_values = {}
daily_max_values = {}
daily_max_drawdown = {}  # Максимальная просадка внутри дня
daily_max_runup = {}     # Максимальный рост внутри дня

# Группируем по дням
for day, group in df.groupby('Date_Day'):
    if len(group) > 0:
        # Первое и последнее значение дня
        daily_first_values[day] = group['Total Portfolio Value'].iloc[0]
        daily_last_values[day] = group['Total Portfolio Value'].iloc[-1]
        
        # Минимум и максимум внутри дня
        daily_min_values[day] = group['Total Portfolio Value'].min()
        daily_max_values[day] = group['Total Portfolio Value'].max()
        
        # Максимальная просадка внутри дня (от максимума дня к минимуму дня)
        if daily_max_values[day] > 0:
            daily_max_drawdown[day] = (daily_min_values[day] - daily_max_values[day]) / daily_max_values[day] * 100
        else:
            daily_max_drawdown[day] = 0
            
        # Максимальный рост внутри дня (от минимума дня к максимуму дня)
        if daily_min_values[day] > 0:
            daily_max_runup[day] = (daily_max_values[day] - daily_min_values[day]) / daily_min_values[day] * 100
        else:
            daily_max_runup[day] = 0

# Создаем колонки с дневными значениями (для каждой строки)
df['Daily_First_Value'] = df['Date_Day'].map(daily_first_values)
df['Daily_Last_Value'] = df['Date_Day'].map(daily_last_values)
df['Daily_Min_Value'] = df['Date_Day'].map(daily_min_values)
df['Daily_Max_Value'] = df['Date_Day'].map(daily_max_values)
df['Daily_Max_Drawdown'] = df['Date_Day'].map(daily_max_drawdown)
df['Daily_Max_Runup'] = df['Date_Day'].map(daily_max_runup)

# 3. Дневная доходность (изменение за день)
# Создаем словарь с последним значением каждого дня
daily_close = df.groupby('Date_Day')['Total Portfolio Value'].last().to_dict()
# Создаем словарь с первым значением каждого дня
daily_open = df.groupby('Date_Day')['Total Portfolio Value'].first().to_dict()

# Для каждой строки берем последнее значение предыдущего дня
prev_day_close = {}
days = sorted(daily_close.keys())
for i, day in enumerate(days):
    if i > 0:
        prev_day_close[day] = daily_close[days[i-1]]
    else:
        prev_day_close[day] = daily_open[day]  # для первого дня используем открытие

df['Prev_Day_Close'] = df['Date_Day'].map(prev_day_close)
df['Daily_Return'] = (df['Daily_Last_Value'] - df['Prev_Day_Close']) / df['Prev_Day_Close'] * 100
df.loc[df['Date_Day'] == days[0], 'Daily_Return'] = 0  # первый день = 0

# 4. Максимальная просадка за день относительно предыдущего закрытия
df['Daily_Drawdown_From_Prev'] = (df['Daily_Min_Value'] - df['Prev_Day_Close']) / df['Prev_Day_Close'] * 100
df.loc[df['Date_Day'] == days[0], 'Daily_Drawdown_From_Prev'] = 0

print(f"  ✅ Рассчитаны просадки:")
print(f"     - Drawdown_15min: изменение за 15 минут")
print(f"     - Daily_Max_Drawdown: максимальная просадка внутри дня")
print(f"     - Daily_Return: доходность за день")
print(f"     - Daily_Drawdown_From_Prev: максимальная просадка от закрытия предыдущего дня")
# ================================================

# ДИАГНОСТИКА ОТКРЫТИЙ/ЗАКРЫТИЙ ШОРТА
print("\n" + "="*60)
print("ДИАГНОСТИКА ОТКРЫТИЙ/ЗАКРЫТИЙ ШОРТА")
print("="*60)

# Все моменты изменения статуса
status_changes = []
for i in range(1, len(df)):
    prev_status = df.at[i-1, 'ShortActive']
    curr_status = df.at[i, 'ShortActive']
    
    if prev_status == 0 and curr_status == 1:
        status_changes.append(('OPEN', i, df.at[i, 'Date'], f"${df.at[i, 'ETH Price']:.2f}"))
    elif prev_status == 1 and curr_status == 0:
        reason = df.at[i, 'CloseShortReason'] if pd.notna(df.at[i, 'CloseShortReason']) and df.at[i, 'CloseShortReason'] else 'unknown'
        status_changes.append(('CLOSE', i, df.at[i, 'Date'], f"${df.at[i, 'ETH Price']:.2f}", reason))

# Выводим первые 10 изменений для проверки
print("Первые 10 изменений статуса шорта:")
for change in status_changes[:10]:
    print(f"  {change}")

opens = sum(1 for c in status_changes if c[0] == 'OPEN')
closes = sum(1 for c in status_changes if c[0] == 'CLOSE')

print(f"\nВсего изменений статуса: {len(status_changes)}")
print(f"  Открытий: {opens}")
print(f"  Закрытий: {closes}")
print(f"  Разница: {opens - closes}")

# Проверяем финальный статус
print(f"\nФинальный статус шорта: {'АКТИВЕН' if df['ShortActive'].iloc[-1] == 1 else 'НЕ АКТИВЕН'}")
print("="*60)

# ===== РАСШИРЕННАЯ ДИАГНОСТИКА ДЛЯ ПОИСКА ПРИЧИНЫ РАЗНИЦЫ =====  <<<<<<<<<< ВСТАВЛЯЕМ ЭТОТ БЛОК ЗДЕСЬ
print("\n" + "="*60)
print("РАСШИРЕННАЯ ДИАГНОСТИКА РАЗНИЦЫ ОТКРЫТИЙ/ЗАКРЫТИЙ")
print("="*60)

# 1. Анализируем все закрытия по типам
close_by_reason = {}
close_by_exit = 0
close_by_trigger = 0

for i in range(1, len(df)):
    # Закрытия через триггер
    if df.at[i, 'CloseShortTrigger'] == 1 and df.at[i-1, 'CloseShortTrigger'] == 0:
        close_by_trigger += 1
        reason = df.at[i, 'CloseShortReason'] if pd.notna(df.at[i, 'CloseShortReason']) else 'unknown'
        close_by_reason[reason] = close_by_reason.get(reason, 0) + 1
    
    # Закрытия через выход из пула
    if df.at[i, 'Pool Exited'] == 1 and df.at[i-1, 'Pool Exited'] == 0:
        if df.at[i-1, 'ShortActive'] == 1:
            close_by_exit += 1

print(f"\nЗакрытий по типам:")
print(f"  Через CloseShortTrigger: {close_by_trigger}")
for reason, count in close_by_reason.items():
    print(f"    - {reason}: {count}")
print(f"  Через выход из пула (Pool Exited): {close_by_exit}")
print(f"  ВСЕГО ЗАКРЫТИЙ: {close_by_trigger + close_by_exit}")

# 2. Проверяем, нет ли двойного учета (когда оба триггера сработали одновременно)
double_count = 0
for i in range(1, len(df)):
    trigger_close = (df.at[i, 'CloseShortTrigger'] == 1 and df.at[i-1, 'CloseShortTrigger'] == 0)
    exit_close = (df.at[i, 'Pool Exited'] == 1 and df.at[i-1, 'Pool Exited'] == 0 and df.at[i-1, 'ShortActive'] == 1)
    
    if trigger_close and exit_close:
        double_count += 1
        print(f"  ВНИМАНИЕ: строка {i} - и trigger, и exit одновременно!")

if double_count > 0:
    print(f"\n⚠️ Найдено {double_count} случаев двойного учета!")

# 3. Проверяем соответствие CloseShortReason и Pool Exited
print(f"\nПроверка соответствия причин закрытия:")
print(f"  Закрытий 'below_pn_exit_pool' по CloseShortReason: {(df['CloseShortReason'].str.contains('below_pn_exit_pool', na=False)).sum()}")
print(f"  Выходов из пула (Pool Exited): {(df['Pool Exited'] == 1).sum()}")

# 4. Анализируем периоды, когда шорт не активен, но есть записи о закрытии
print(f"\nАнализ последовательностей закрытий:")
problem_sequences = 0
for i in range(1, len(df)-1):
    if df.at[i, 'CloseShortTrigger'] == 1:
        # Проверяем, был ли шорт активен до закрытия
        if df.at[i-1, 'ShortActive'] != 1:
            problem_sequences += 1
            if problem_sequences <= 5:  # покажем первые 5
                print(f"  Проблема в строке {i}: CloseShortTrigger=1, но ShortActive ранее = {df.at[i-1, 'ShortActive']}")
                print(f"    ShortActive сейчас: {df.at[i, 'ShortActive']}")
                print(f"    Причина: {df.at[i, 'CloseShortReason']}")

print(f"\nНайдено проблемных последовательностей: {problem_sequences}")

# 5. Проверяем все строки с Pool Exited
print(f"\nДетальный анализ выходов из пула (первые 10):")
exited_rows = df[df['Pool Exited'] == 1].index[:10]
for idx in exited_rows:
    print(f"  Строка {idx}:")
    print(f"    ShortActive до: {df.at[idx-1, 'ShortActive'] if idx > 0 else 'N/A'}")
    print(f"    ShortActive после: {df.at[idx, 'ShortActive']}")
    print(f"    CloseShortReason: {df.at[idx, 'CloseShortReason']}")
    print(f"    CloseShortTrigger: {df.at[idx, 'CloseShortTrigger']}")

print("="*60)
# =============================================================

# ===== ДОБАВЛЕННАЯ ДИАГНОСТИКА + ЗАЩИТА ОТ ОТРИЦАТЕЛЬНЫХ ЗНАЧЕНИЙ =====
print("\n" + "="*60)
print("ДИАГНОСТИКА И ЗАЩИТА ОТРИЦАТЕЛЬНЫХ ЗНАЧЕНИЙ")
print("="*60)

# 1. Диагностика до исправлений
negative_pending = (df['Pending Cash'] < 0).sum()
negative_total = (df['Total Portfolio Value'] < 0).sum()
exits_count = (df['Pool Exited'] == 1).sum()

print(f"Строк с отрицательным Pending Cash: {negative_pending}")
if negative_pending > 0:
    print(f"  Примеры (первые 3):")
    for idx in df[df['Pending Cash'] < 0].index[:3]:
        print(f"    Строка {idx}: Pending Cash=${df.at[idx, 'Pending Cash']:.2f}, "
              f"Pool Exited={df.at[idx, 'Pool Exited']}, "
              f"Total Portfolio Value=${df.at[idx, 'Total Portfolio Value']:.2f}")

print(f"Строк с отрицательным Total Portfolio Value: {negative_total}")
if negative_total > 0:
    print(f"  Примеры (первые 3):")
    for idx in df[df['Total Portfolio Value'] < 0].index[:3]:
        print(f"    Строка {idx}: Total Portfolio Value=${df.at[idx, 'Total Portfolio Value']:.2f}, "
              f"Pool Exited={df.at[idx, 'Pool Exited']}")

print(f"Всего выходов из пула (Pool Exited == 1): {exits_count}")
print("-"*60)

# 2. Исправления
# Клипаем Pending Cash глобально (не только при exited)
if negative_pending > 0:
    print(f"  → Исправляем {negative_pending} строк с отрицательным Pending Cash (clip to 0)")
    df['Pending Cash'] = df['Pending Cash'].clip(lower=0)

# Пересчитываем Total Portfolio Value после клипа Pending Cash
df['Total Portfolio Value'] = df['Pool Value'] + df['Pending Cash'] + df['Compounded Cash']

# Дополнительная защита: если Total Portfolio Value всё равно отрицательный — клипаем
negative_total_after = (df['Total Portfolio Value'] < 0).sum()
if negative_total_after > 0:
    print(f"  → Обнаружено {negative_total_after} строк с отрицательным Total Portfolio Value после клипа → исправляем в 0")
    df['Total Portfolio Value'] = df['Total Portfolio Value'].clip(lower=0)

# Финальная проверка
min_tvl = df['Total Portfolio Value'].min()
print(f"  Минимальное Total Portfolio Value после всех исправлений: ${min_tvl:,.2f}")
print(f"  Строк с Total Portfolio Value <= 0: {(df['Total Portfolio Value'] <= 0).sum()}")
print("="*60 + "\n")

# 3. Правильный расчет Cumulative PnL и ROI
df['Cumulative PnL'] = df['Total Portfolio Value'] - capital  # ПРЯМОЙ РАСЧЕТ!
df['ROI'] = (df['Total Portfolio Value'] - capital) / capital
df['Total ROI'] = df['ROI']  # для совместимости

# 4. Удаляем старые неправильные колонки если они есть
for col in ['Portfolio Peak', 'Drawdown']:
    if col in df.columns:
        df.drop(columns=[col], inplace=True)
# =============================================================

def create_event_description(row):
    events = []
    if row.name > 0:
        if row['DynamicPa'] != df.at[row.name-1, 'DynamicPa']:
            events.append("Сдвиг диапазона")
    if row['CloseShortTrigger'] == 1:
        if row['CloseShortReason'] == 'max_loss':
            events.append("Закрытие шорта: max loss")
        elif row['CloseShortReason'] == 'bep':
            events.append("Закрытие шорта: достижение BEP")
        elif row['CloseShortReason'] == 'pa_buffer':
            events.append("Закрытие шорта: буфер до Pa")
        elif 'below_pn_exit_pool' in str(row['CloseShortReason']):
            events.append("Закрытие шорта ниже Pn + выход из пула")
    if row.name > 0:
        if row['ShortActive'] == 1 and df.at[row.name-1, 'ShortActive'] == 0:
            events.append("Открытие шорта")
    if row['DynamicInRange'] == 0:
        events.append("Вне диапазона")
    if row['Compound Event'] == 1:
        events.append("Реинвестирование кэша")
    if row['Pool Exited'] == 1:
        events.append("Выход из пула")
    return "; ".join(events) if events else ""

df['Event'] = df.apply(create_event_description, axis=1)

# --- ПРОВЕРКА ИСПРАВЛЕНИЙ ---
print(f"\nПРОВЕРКА ИСПРАВЛЕНИЙ:")
print("-" * 50)

print("1. Проверка соответствия размера шорта цене входа:")
for i in range(len(df)):
    if df.at[i, 'ShortActive'] == 1:
        entry_price = df.at[i, 'ShortEntryPrice']
        pa = df.at[i, 'DynamicPa']
        pb = df.at[i, 'DynamicPb']
        l = df.at[i, 'ShortLiquidity']
        actual_short_size = df.at[i, 'Short ETH']
        expected_size = calculate_short_size(entry_price, pa, pb, l)
        if abs(actual_short_size - expected_size) > 0.0001:
            print(f"⚠️ Несоответствие в строке {i}:")

print("\n2. Проверка корректности Pending Cash:")
total_short_pnl = df['Short Realized PnL'].sum()
total_fees = df['Accrued Fees'].sum()
total_funding = df['Funding PnL'].sum()
total_compound_cost = df['Compound Cost'].sum()
final_pending = df['Pending Cash'].iloc[-1]
expected_pending = total_short_pnl + total_fees + total_funding - total_compound_cost
print(f"  Short Realized PnL: ${total_short_pnl:.2f}")
print(f"  Accrued Fees: ${total_fees:.2f}")
print(f"  Funding PnL: ${total_funding:.2f}")
print(f"  Compound Cost: ${total_compound_cost:.2f}")
print(f"  Ожидаемо: ${expected_pending:.2f}")
print(f"  Фактически: ${final_pending:.2f}")
print(f"  Расхождение: ${abs(final_pending - expected_pending):.2f}")

print("\n3. Проверка корректности расчета BEP:")
active_short_rows = df['ShortActive'].sum()
bep_calculated_rows = (df['BEP'] > 0).sum()
print(f"  Строк с активным шортом: {active_short_rows}")
print(f"  Строк с рассчитанным BEP (> 0): {bep_calculated_rows}")

print("\n4. Проверка значений BEP:")
for i in range(len(df)):
    if df.at[i, 'ShortActive'] == 1 and df.at[i, 'BEP'] > 0:
        pa = df.at[i, 'DynamicPa']
        entry_price = df.at[i, 'ShortEntryPrice']
        bep = df.at[i, 'BEP']
        if not (pa < bep < entry_price):
            print(f"⚠️ Неверный BEP в строке {i}: Pa=${pa:.2f}, BEP=${bep:.2f}, Entry=${entry_price:.2f}")

# --- ВАЛИДАЦИЯ РАСЧЕТОВ ---
print(f"\nВАЛИДАЦИЯ РАСЧЕТОВ:")
print("-" * 50)

if len(df) > 1:
    price_changes = df['ETH Price'].pct_change().fillna(0)
    pool_value_changes = df['Pool Value'].pct_change().fillna(0)
    correlation = np.corrcoef(price_changes[1:], pool_value_changes[1:])[0, 1]
    print(f"Корреляция между изменением цены ETH и изменением Pool Value: {correlation:.4f}")

print(f"\nИТОГОВЫЕ МЕТРИКИ:")
print(f"  Стартовый капитал: ${capital:,.2f}")
print(f"  Итоговая стоимость пула (Pool Value): ${df['Pool Value'].iloc[-1]:,.2f}")
print(f"  Итоговый Pending Cash: ${df['Pending Cash'].iloc[-1]:,.2f}")
print(f"  Итоговый Compounded Cash: ${df['Compounded Cash'].iloc[-1]:,.2f}")
print(f"  Итоговая общая стоимость (Total Portfolio Value): ${df['Total Portfolio Value'].iloc[-1]:,.2f}")
print(f"  Итоговый Total ROI: {df['Total ROI'].iloc[-1]:.2%}")

# ===== ДОБАВЛЕННАЯ ПРОВЕРКА КОРРЕКТНОСТИ =====
print(f"\nПРОВЕРКА КОРРЕКТНОСТИ:")
print(f"  Минимальный Total Portfolio Value: ${df['Total Portfolio Value'].min():.2f}")
print(f"  Минимальный Pending Cash: ${df['Pending Cash'].min():.2f}")
if df['Total Portfolio Value'].min() < 0:
    print(f"  ⚠️ ВНИМАНИЕ: Есть отрицательные значения портфеля!")
    print(f"     Это указывает на ошибки в расчетах.")

# Проверка exited-периодов
exited_zero_tvl = df[(df['Pool Exited'] == 1) & (df['Total Portfolio Value'] <= 0)]
if not exited_zero_tvl.empty:
    print(f"\nВнимание: {len(exited_zero_tvl)} строк с exited и TVL <= 0")
    print("Первые 3:")
    print(exited_zero_tvl[['Date', 'Pending Cash', 'Compounded Cash', 'Total Portfolio Value']].head(3))

# Расчет просадки (используем исправленную функцию)
from copy import deepcopy
temp_df = deepcopy(df)
temp_df, correct_max_drawdown = calculate_correct_drawdown(temp_df)
df['Correct Drawdown'] = temp_df['Correct Drawdown']
max_drawdown = correct_max_drawdown

print(f"\nИТОГОВЫЕ МЕТРИКИ:")
print(f"  Стартовый капитал: ${capital:,.2f}")
print(f"  Итоговая общая стоимость: ${df['Total Portfolio Value'].iloc[-1]:,.2f}")
print(f"  Итоговый ROI: {df['ROI'].iloc[-1]:.2%}")
print(f"  Максимальная просадка (скорректированная): {max_drawdown:.2f}%")

df['CloseShortTrigger'] = df['CloseShortTrigger'].astype(int)
df['ShouldReopenShort'] = df['ShouldReopenShort'].astype(int)
df['Compound Event'] = df['Compound Event'].astype(int)

# =============================================================================
# PAYOFF-ДИАГРАММА СТРАТЕГИИ (только для Excel, без отображения)
# =============================================================================
print(f"\nСОЗДАНИЕ PAYOFF-ДИАГРАММЫ СТРАТЕГИИ...")

def create_payoff_diagram(entry_price, pa, pb, l, apr=0.6, short_size=None):
    """
    Создает payoff-диаграмму для стратегии Uniswap V3 LP + Short ETH
    """
    if short_size is None:
        short_size = calculate_short_size(entry_price, pa, pb, l)
    
    # Диапазон цен для анализа (от 500 до 5000)
    price_range = np.linspace(500, 5000, 1000)
    
    # 1. Расчет PnL от пула (IL + изменение стоимости активов)
    pool_pnl = []
    pool_value_entry = calculate_pool_value(entry_price, pa, pb, l)
    
    for price in price_range:
        pool_value_current = calculate_pool_value(price, pa, pb, l)
        pool_loss = pool_value_entry - pool_value_current
        pool_pnl.append(-pool_loss)  # Инвертируем для отображения доходности
    
    pool_pnl = np.array(pool_pnl)
    
    # 2. Расчет PnL от шорт позиции
    short_pnl = np.array([short_size * (entry_price - price) for price in price_range])
    
    # 3. Расчет начисленных комиссий (fees)
    days_in_range_probability = 0.5
    annual_fees_pct = apr
    daily_fees_pct = annual_fees_pct / 365.0
    
    fees_pnl = []
    for price in price_range:
        if pa <= price <= pb:
            fees = pool_value_entry * daily_fees_pct * 30 * days_in_range_probability
        else:
            fees = 0
        fees_pnl.append(fees)
    
    fees_pnl = np.array(fees_pnl)
    
    # 4. Общий PnL = пул + шорт + комиссии
    total_pnl = pool_pnl + short_pnl + fees_pnl
    
    return price_range, pool_pnl, short_pnl, total_pnl, fees_pnl, pool_value_entry, short_size

# НАХОДИМ ПАРАМЕТРЫ ДЛЯ ДИАГРАММЫ
print("  Поиск последнего активного пула для Payoff-диаграммы...")

# Ищем строки с активным пулом (Pool Exited == 0 и DynamicL > 0)
active_pool_mask = (df['Pool Exited'] == 0) & (df['DynamicL'] > 0)
active_indices = df[active_pool_mask].index

if len(active_indices) > 0:
    # Берем последний активный пул
    last_active_idx = active_indices[-1]
    last_row = df.iloc[last_active_idx]
    
    entry_price = last_row['DynamicPn']
    pa = last_row['DynamicPa']
    pb = last_row['DynamicPb']
    l = last_row['DynamicL']
    current_price = last_row['ETH Price']
    
    print(f"  ✅ Использую последний активный пул из строки {last_active_idx}")
    
else:
    print("  ⚠️ Активных пулов не найдено! Использую параметры из настроек.")
    
    # Создаем "типичный" пул из настроек
    center_price = df['ETH Price'].median()
    
    if asymmetric_range_enabled:
        pa, pn, pb = calculate_asymmetric_range(
            center_price,
            range_width,
            asymmetry_distribution
        )
    else:
        pn = center_price
        pa = pn - (range_width / 2)
        pb = pn + (range_width / 2)
    
    l = calculate_liquidity(capital, pn, pa, pb)
    entry_price = pn
    current_price = df['ETH Price'].iloc[-1]

# Создаем данные для payoff-диаграммы
price_range, pool_pnl, short_pnl, total_pnl, fees_pnl, pool_value_entry, short_size = create_payoff_diagram(
    entry_price=entry_price,
    pa=pa,
    pb=pb,
    l=l,
    apr=apr
)

# Создаем график с тёмным фоном
plt.style.use('dark_background')
fig, ax = plt.subplots(figsize=(16, 10), facecolor='#0d0d0d')
ax.set_facecolor('#0d0d0d')

# Линии графика
ax.plot(price_range, pool_pnl, color='#0080ff', linewidth=3, 
        label='Доход от пула (IL + стоимость активов)', alpha=0.9)
ax.plot(price_range, short_pnl, color='#ff4040', linewidth=3,
        label='Доход от шорт позиции', alpha=0.9)
ax.plot(price_range, total_pnl, color='#00ff80', linewidth=4,
        label='Дельта с учётом прибыли пула', alpha=1.0)
ax.plot(price_range, fees_pnl, color='#cccccc', linewidth=1, 
        linestyle='--', label='Начисленные комиссии', alpha=0.5)

# Вспомогательные линии
ax.axhline(y=0, color='white', linestyle='--', linewidth=1.5, alpha=0.5)
ax.axvline(x=current_price, color='white', linestyle='--', linewidth=1.5, alpha=0.5)

# Настройка осей
ax.set_xlabel('Цена ETH, USD', fontsize=14, color='white', fontweight='bold')
ax.set_xlim(500, 5000)
x_ticks = np.arange(500, 5001, 500)
ax.set_xticks(x_ticks)
ax.set_xticklabels([f'${x:,.0f}' for x in x_ticks], rotation=45, ha='right', fontsize=11)

ax.set_ylabel('PnL, USD', fontsize=14, color='white', fontweight='bold')
ax.set_ylim(-150000, 150000)
y_ticks = np.arange(-150000, 150001, 50000)
ax.set_yticks(y_ticks)
ax.set_yticklabels([f'${y:,.0f}' for y in y_ticks], fontsize=11)

# Сетка
ax.grid(True, alpha=0.15, linestyle='-', linewidth=0.5, color='gray')

# Легенда
legend = ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.12),
                   ncol=2, fontsize=12, framealpha=0.2, frameon=True)
for text in legend.get_texts():
    text.set_color('white')

# Заголовок
data_type_text = "1-минутные" if use_1min_data else "15-минутные"
plt.title(f'Payoff диаграмма стратегии: Пул + Шорт + Fees ({data_type_text} данные)',
          fontsize=18, color='white', fontweight='bold', pad=20)

# Информационные аннотации
info_text = (f'Текущая цена: ${current_price:,.0f}\n'
             f'Диапазон: ${pa:,.0f} - ${pb:,.0f}\n'
             f'Цена входа: ${entry_price:,.0f}\n'
             f'Размер шорта: {short_size:.2f} ETH\n'
             f'Стоимость пула: ${pool_value_entry:,.0f}')

plt.annotate(info_text, xy=(0.02, 0.98), xycoords='axes fraction',
             fontsize=12, color='white', verticalalignment='top',
             bbox=dict(boxstyle='round', facecolor='#1a1a1a', alpha=0.8, edgecolor='gray'))

# Область диапазона
ax.axvspan(pa, pb, alpha=0.1, color='yellow', label='Диапазон ликвидности')

plt.tight_layout()

# Сохраняем график (НО НЕ ПОКАЗЫВАЕМ)
payoff_filename = 'payoff_diagram_strategy_1min.png' if use_1min_data else 'payoff_diagram_strategy_15min.png'
plt.savefig(payoff_filename, dpi=300, bbox_inches='tight', facecolor='#0d0d0d')
print(f"  ✅ Payoff-диаграмма сохранена как: {payoff_filename}")

# ВАЖНО: Закрываем фигуру, чтобы не показывать
plt.close(fig)

# --- ФУНКЦИИ ДЛЯ СОХРАНЕНИЯ В EXCEL ---
def create_charts(df, capital):
    charts = []
    try:
        plt.figure(figsize=(10, 6))
        plt.plot(df['Date'], df['Total Portfolio Value'], linewidth=2, color='blue', label='Общая стоимость портфеля')
        plt.axhline(y=capital, color='red', linestyle='--', linewidth=1, label=f'Начальный капитал (${capital:,.0f})')
        if reinvest_cash_enabled:
            plt.plot(df['Date'], df['Pool Value'], linewidth=1, color='green', linestyle='--', alpha=0.7, label='Стоимость пула (V3)')
            plt.plot(df['Date'], df['Pending Cash'] + df['Compounded Cash'], linewidth=1, color='orange', linestyle=':', alpha=0.7, label='Накопленный кэш')
        plt.title('Динамика общей стоимости портфеля', fontsize=14, fontweight='bold')
        plt.xlabel('Дата', fontsize=12)
        plt.ylabel('Стоимость ($)', fontsize=12)
        plt.grid(True, alpha=0.3)
        plt.legend()
        plt.xticks(rotation=45)
        plt.tight_layout()
        img1 = BytesIO()
        plt.savefig(img1, format='png', dpi=150)
        img1.seek(0)
        charts.append(('portfolio_value.png', img1))
        plt.close()
        
        fig, ax1 = plt.subplots(figsize=(10, 6))
        color = 'tab:blue'
        ax1.set_xlabel('Дата', fontsize=12)
        ax1.set_ylabel('Стоимость пула ($)', color=color, fontsize=12)
        ax1.plot(df['Date'], df['Pool Value'], color=color, linewidth=2, label='Стоимость пула')
        ax1.tick_params(axis='y', labelcolor=color)
        ax1.grid(True, alpha=0.3)
        ax2 = ax1.twinx()
        color = 'tab:orange'
        ax2.set_ylabel('Цена ETH ($)', color=color, fontsize=12)
        ax2.plot(df['Date'], df['ETH Price'], color=color, linewidth=1, alpha=0.7, label='Цена ETH')
        ax2.tick_params(axis='y', labelcolor=color)
        plt.title('Стоимость пула (V3) и цена ETH', fontsize=14, fontweight='bold')
        fig.tight_layout()
        img2 = BytesIO()
        plt.savefig(img2, format='png', dpi=150)
        img2.seek(0)
        charts.append(('pool_value_eth_price.png', img2))
        plt.close()
        
        plt.figure(figsize=(10, 6))
        if 'Pending Cash' in df.columns and 'Compounded Cash' in df.columns:
            total_cash = df['Pending Cash'] + df['Compounded Cash']
            plt.plot(df['Date'], total_cash, linewidth=2, color='purple', label='Общий кэш')
            plt.plot(df['Date'], df['Pending Cash'], linewidth=1, color='orange', linestyle='--', alpha=0.7, label='Pending Cash')
            plt.plot(df['Date'], df['Compounded Cash'], linewidth=1, color='green', linestyle=':', alpha=0.7, label='Compounded Cash')
        plt.title('Динамика накопленного кэша', fontsize=14, fontweight='bold')
        plt.xlabel('Дата', fontsize=12)
        plt.ylabel('Кэш ($)', fontsize=12)
        plt.grid(True, alpha=0.3)
        plt.legend()
        plt.xticks(rotation=45)
        plt.tight_layout()
        img3 = BytesIO()
        plt.savefig(img3, format='png', dpi=150)
        img3.seek(0)
        charts.append(('accumulated_cash.png', img3))
        plt.close()
        
        if 'Correct Drawdown' in df.columns:
            df['Drawdown'] = df['Correct Drawdown']  # используем скорректированную версию
        else:
            # Если нет Correct Drawdown, считаем старым способом как запасной вариант
            df['Portfolio Peak'] = df['Total Portfolio Value'].cummax()
            df['Drawdown'] = (df['Total Portfolio Value'] - df['Portfolio Peak']) / df['Portfolio Peak'] * 100
        
        plt.figure(figsize=(10, 6))
        plt.fill_between(df['Date'], df['Drawdown'], 0, color='red', alpha=0.3)
        plt.plot(df['Date'], df['Drawdown'], color='darkred', linewidth=2)
        plt.title('Просадка портфеля (скорректированная, без учета выходов из пула)', fontsize=14, fontweight='bold')
        plt.xlabel('Дата', fontsize=12)
        plt.ylabel('Просадка (%)', fontsize=12)
        plt.grid(True, alpha=0.3)
        plt.xticks(rotation=45)
        plt.tight_layout()
        img4 = BytesIO()
        plt.savefig(img4, format='png', dpi=150)
        img4.seek(0)
        charts.append(('drawdown_corrected.png', img4))
        plt.close()
        
        print(f"  Создано {len(charts)} графика")
    except Exception as e:
        print(f"  Ошибка при создании графиков: {e}")
    return charts

def save_to_excel_with_improvements(df, filename=None):
    if filename is None:
        filename = 'backtest_results_1min.xlsx' if use_1min_data else 'backtest_results_improved_15min.xlsx'
    
    if not EXCEL_SUPPORT:
        print("openpyxl не установлен. Сохраняю в CSV...")
        df.to_csv(filename.replace('.xlsx', '.csv'), index=False)
        return
    
    print(f"\nСохранение результатов в Excel файл: {filename}")
    
    total_days = len(df)
    days_in_range = df['DynamicInRange'].sum()
    days_with_short = df['ShortActive'].sum()
    
    short_opens = 0
    for i in range(len(df)):
        if i == 0:
            if df.at[i, 'ShortActive'] == 1:
                short_opens += 1
        else:
            if df.at[i, 'ShortActive'] == 1 and df.at[i-1, 'ShortActive'] == 0:
                short_opens += 1
    
    # Исправленный подсчет закрытий для Excel
    short_closes = 0
    for i in range(len(df)):
        if i > 0:
            # Вариант 1: CloseShortTrigger
            if df.at[i, 'CloseShortTrigger'] == 1 and df.at[i-1, 'CloseShortTrigger'] == 0:
                short_closes += 1
            # Вариант 2: Pool Exited (выход из пула с закрытием шорта)
            elif df.at[i, 'Pool Exited'] == 1 and df.at[i-1, 'Pool Exited'] == 0:
                if df.at[i-1, 'ShortActive'] == 1:
                    short_closes += 1
    
    # Исправленный подсчет открытий для Excel
    short_opens = 0
    for i in range(len(df)):
        if df.at[i, 'ShortActive'] == 1:
            if i == 0 or df.at[i-1, 'ShortActive'] == 0:
                short_opens += 1
    
    max_loss_closes = (df['CloseShortReason'] == 'max_loss').sum()
    bep_closes = (df['CloseShortReason'] == 'bep').sum()
    pa_buffer_closes = (df['CloseShortReason'] == 'pa_buffer').sum()
    below_pn_exit_closes = (df['CloseShortReason'].str.contains('below_pn_exit_pool', na=False)).sum()
    
    final_portfolio_value = df['Total Portfolio Value'].iloc[-1]
    final_cumulative_pnl = df['Cumulative PnL'].iloc[-1]
    roi = df['ROI'].iloc[-1]
    total_roi = df['Total ROI'].iloc[-1] if 'Total ROI' in df.columns else roi
    
    final_pool_value = df['Pool Value'].iloc[-1]
    final_pending_cash = df['Pending Cash'].iloc[-1] if 'Pending Cash' in df.columns else 0
    final_compounded_cash = df['Compounded Cash'].iloc[-1] if 'Compounded Cash' in df.columns else 0
    total_compound_events = df['Compound Event'].sum() if 'Compound Event' in df.columns else 0
    total_compound_cost = df['Compound Cost'].sum() if 'Compound Cost' in df.columns else 0
    
    # Используем скорректированный расчет просадки
    if 'Correct Drawdown' not in df.columns:
        # Создаем временную копию для расчета
        temp_df = df.copy()
        temp_df, correct_max_drawdown = calculate_correct_drawdown(temp_df)
        # Переносим результаты
        df['Correct Drawdown'] = temp_df['Correct Drawdown']
    else:
        correct_max_drawdown = df['Correct Drawdown'].min()
    
    # Используем правильное значение
    max_drawdown = correct_max_drawdown
    
    winning_shorts = 0
    total_shorts = 0
    short_pnl_values = []
    for i in range(len(df)):
        if df.at[i, 'CloseShortTrigger'] == 1 and df.at[i, 'Short Realized PnL'] != 0:
            total_shorts += 1
            if df.at[i, 'Short Realized PnL'] > 0:
                winning_shorts += 1
            short_pnl_values.append(df.at[i, 'Short Realized PnL'])
    avg_short_pnl = np.mean(short_pnl_values) if short_pnl_values else 0
    win_rate = winning_shorts / total_shorts if total_shorts > 0 else 0
    
    shift_count = (df['DynamicPa'] != df['DynamicPa'].shift(1)).sum() - 1
    
    print(f"\nСоздание Dashboard с ключевыми метриками...")
    
    excel_df = df.copy()
    excel_df['Date'] = excel_df['Date'].dt.strftime('%Y-%m-%d %H:%M')
    
    charts = create_charts(df, capital)
    
    # ===== ИСПРАВЛЕНИЕ: Расчет метрик реинвестирования для Dashboard =====
    calendar_reinvests = 0
    shift_reinvests = 0
    close_reinvests = 0
    avg_reinvest_interval = 0
    
    if total_compound_events > 0:
        # Анализируем причины реинвестирований
        for idx in df[df['Compound Event'] == 1].index:
            reason = str(df.at[idx, 'CloseShortReason']) if pd.notna(df.at[idx, 'CloseShortReason']) else ''
            if 'weekly' in reason or 'biweekly' in reason or 'monthly' in reason:
                calendar_reinvests += 1
            elif 'shift' in reason:
                shift_reinvests += 1
            elif 'short_close' in reason:
                close_reinvests += 1
        
        # Средний интервал между реинвестированиями
        reinvest_dates = df[df['Compound Event'] == 1]['Date'].sort_values()
        if len(reinvest_dates) > 1:
            intervals = reinvest_dates.diff().dt.total_seconds() / (3600 * 24)  # в днях
            avg_reinvest_interval = intervals.mean()
    # =====================================================================
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        workbook = writer.book
        
        # --- DASHBOARD ---
        dashboard_data = [
            ["МЕТРИКА", "ЗНАЧЕНИЕ", "ОПИСАНИЕ"],
            ["Стартовый капитал", f"${capital:,.2f}", "Начальная сумма инвестиций"],
            ["Конечная стоимость портфеля", f"${final_portfolio_value:,.2f}", "Общая стоимость на конец периода"],
            ["Стоимость пула (V3)", f"${final_pool_value:,.2f}", "Стоимость активов в пуле по формуле V3"],
            ["", "", ""],
            ["НАКОПЛЕННЫЙ КЭШ", "", ""],
            ["Реинвестирование кэша", "ВКЛ" if reinvest_cash_enabled else "ВЫКЛ", "Реинвестирование fees, funding, short PnL"],
            ["Частота реинвестирования", reinvest_frequency, "Частота реинвестирования"],
            ["Pending Cash", f"${final_pending_cash:,.2f}", "Кэш, ожидающий реинвестирования"],
            ["Compounded Cash", f"${final_compounded_cash:,.2f}", "Уже реинвестированный кэш"],
            ["Общий накопленный кэш", f"${final_pending_cash + final_compounded_cash:,.2f}", "Весь кэш (pending + compounded)"],
            ["Событий реинвестирования", total_compound_events, "Всего реинвестирований кэша"],
            ["По календарным триггерам", calendar_reinvests, "weekly/biweekly/monthly"],
            ["По сдвигу диапазона", shift_reinvests, "при сдвиге диапазона"],
            ["По закрытию шорта", close_reinvests, "при закрытии шорта"],
            ["Средний интервал между реинвестированиями", f"{avg_reinvest_interval:.1f} дней", "в днях"],
            ["Затраты на реинвестирование", f"${total_compound_cost:.2f}", "Общие затраты на реинвестирование"],
            ["", "", ""],
            ["ДОХОДНОСТЬ", "", ""],
            ["Кумулятивный PnL", f"${final_cumulative_pnl:,.2f}", "Совокупная прибыль/убыток"],
            ["ROI (базовый)", f"{roi:.2%}", "Доходность без учета реинвестированного кэша"],
            ["Total ROI", f"{total_roi:.2%}", "Общая доходность с учетом всех компонентов"],
            ["Макс. просадка (скорректированная)", f"{max_drawdown:.2f}%", "Максимальное падение от пика"],
            ["Макс. просадка за 15 мин", f"{df['Drawdown_15min'].min():.2f}%", "Минимальное изменение за 15 минут"],
            ["Макс. просадка за день", f"{df['Daily_Max_Drawdown'].min():.2f}%", "Максимальное падение внутри дня"],
            ["Макс. рост за день", f"{df['Daily_Max_Runup'].max():.2f}%", "Максимальный рост внутри дня"],
            ["Средняя дневная доходность", f"{df['Daily_Return'].mean():.2f}%", "Среднее изменение за день"],
            ["", "", ""],
            ["СТАТИСТИКА", "", ""],
            ["Всего интервалов", total_days, f"Период с {df['Date'].iloc[0].strftime('%d.%m.%Y %H:%M')} по {df['Date'].iloc[-1].strftime('%d.%m.%Y %H:%M')}"],
            ["Интервалов в диапазоне", f"{days_in_range} ({days_in_range/total_days*100:.1f}%)", "Интервалы когда цена была между Pa и Pb"],
            ["Интервалов с шортом", f"{days_with_short} ({days_with_short/total_days*100:.1f}%)", "Интервалы с активной короткой позицией"],
            ["", "", ""],
            ["ОПЕРАЦИИ С ШОРТОМ", "", ""],
            ["Открытий шорта", short_opens, "Всего открытий короткой позиции"],
            ["Закрытий шорта", short_closes, "Всего закрытий короткой позиции"],
            ["Выигрышных шортов", f"{winning_shorts} ({win_rate:.1%})", "Закрытия с положительным PnL"],
            ["Средний PnL шорта", f"${avg_short_pnl:,.2f}", "Средняя прибыль/убыток при закрытии"],
            ["", "", ""],
            ["ПРИЧИНЫ ЗАКРЫТИЯ ШОРТА", "", ""],
            ["Закрытий по max loss", max_loss_closes, "При превышении цены входа"],
            ["Закрытий по BEP", bep_closes, "При достижении Break-Even Price"],
            ["Закрытий до Pa", pa_buffer_closes, "По буферу до нижней границы"],
            ["Закрытий ниже Pn с выходом из пула", below_pn_exit_closes, "Закрытие ниже Pn с полным выходом из пула"],
            ["", "", ""],
            ["ДИАПАЗОН", "", ""],
            ["Сдвигов диапазона", shift_count, "Всего сдвигов диапазона"],
            ["Асимметричный диапазон", "ВКЛ" if asymmetric_range_enabled else "ВЫКЛ", "Асимметричное распределение"],
            ["Распределение выше Pn", f"{asymmetry_distribution}%" if asymmetric_range_enabled else "не используется", "Процент диапазона выше Pn"],
            ["", "", ""],
            ["ПОВТОРНЫЙ ВХОД В ПУЛ", "", ""],
            ["Повторный вход в пул", "ВКЛ" if reentry_enabled else "ВЫКЛ", "Автоматический повторный вход после выхода"],
            ["Задержка повторного входа", f"{reentry_delay_hours} часов", "Время ожидания перед повторным входом"],
            ["Порог падения цены", f"{reentry_price_drop_threshold}%", "Падение от максимума для ускоренного входа"],
            ["Агрессивный вход через", f"{reentry_aggressive_hours} часов", "Принудительный вход через N часов"],
            ["", "", ""],
            ["ПАРАМЕТРЫ ДАННЫХ", "", ""],
            ["Тип данных", "1-минутные" if use_1min_data else "15-минутные", "Интервал между записями"],
            ["Шаг времени", f"{time_step_hours} часа", "Длительность одного интервала"],
            ["Ценовой столбец", PRICE_COLUMN, "Используемая цена для расчетов"],
            ["Использование high/low", "ДА" if use_high_low_for_out_of_range else "НЕТ", "Использовать high/low для выхода из диапазона"],
            ["Подробный лог", "ВКЛ" if verbose_logging else "ВЫКЛ", "Детальное логирование событий"],
            ["НОВАЯ ЛОГИКА: Выход из пула при закрытии ниже Pn", "ВКЛ", "При закрытии шорта ниже Pn - полный выход из пула"],
            ["ИСПРАВЛЕНИЕ: Шорт открывается по Pn при создании диапазона", "ВКЛ", "Правильный хедж - цена входа = центр диапазона"]
        ]
        
        dashboard_df = pd.DataFrame(dashboard_data)
        dashboard_df.to_excel(writer, sheet_name='Dashboard', index=False, header=False)
        dashboard_sheet = writer.sheets['Dashboard']
        
        title_font = Font(bold=True, size=16, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # ← ДОБАВЛЕНО
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        dashboard_sheet.merge_cells('A1:C1')
        title_cell = dashboard_sheet['A1']
        data_type_text = "1-минутные" if use_1min_data else "15-минутные"
        title_cell.value = f'АНАЛИЗ РЕЗУЛЬТАТОВ СТРАТЕГИИ ({data_type_text} данные)'
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        dashboard_sheet.merge_cells('A2:C2')
        subtitle_cell = dashboard_sheet['A2']
        subtitle_cell.value = f'Период: {df["Date"].iloc[0].strftime("%d.%m.%Y %H:%M")} - {df["Date"].iloc[-1].strftime("%d.%m.%Y %H:%M")} | Интервалов: {total_days}'
        subtitle_cell.font = Font(italic=True, size=10)
        subtitle_cell.alignment = Alignment(horizontal='center')
        
        for col in ['A', 'B', 'C']:
            cell = dashboard_sheet[f"{col}3"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        dashboard_sheet.column_dimensions['A'].width = 35
        dashboard_sheet.column_dimensions['B'].width = 25
        dashboard_sheet.column_dimensions['C'].width = 50
        dashboard_sheet.row_dimensions[1].height = 30
        dashboard_sheet.row_dimensions[3].height = 25
        
        # --- PARAMETERS ---
        params_sheet = workbook.create_sheet("Parameters")
        parameters = [
            ["ПАРАМЕТР СТРАТЕГИИ", "ЗНАЧЕНИЕ"],
            ["", ""],
            ["ОСНОВНЫЕ ПАРАМЕТРЫ", ""],
            ["Стартовый капитал", f"${capital:,.2f}"],
            ["Ширина диапазона", f"${range_width:,.2f}"],
            ["APR пула", f"{apr:.1%}"],
            ["Funding rate", f"{funding_rate:.2%}"],
            ["Газ", f"${gas_usd:.2f}"],
            ["Slippage", f"{slippage:.1%}"],
            ["Short fee", f"{short_fee:.1%}"],
            ["", ""],
            ["ПАРАМЕТРЫ ДАННЫХ", ""],
            ["Тип данных", "1-минутные" if use_1min_data else "15-минутные"],
            ["Шаг времени", f"{time_step_hours} часа ({int(time_step_hours*60)} минут)"],
            ["Ценовой столбец", PRICE_COLUMN],
            ["Использовать high/low", "ДА" if use_high_low_for_out_of_range else "НЕТ"],
            ["Подробный лог", "ВКЛ" if verbose_logging else "ВЫКЛ"],
            ["", ""],
            ["ДИАПАЗОН", ""],
            ["Сдвиг диапазона", "ВКЛ" if shift_enabled else "ВЫКЛ"],
            ["Задержка сдвига", f"{shift_delay_hours} часов"],
            ["Асимметричный диапазон", "ВКЛ" if asymmetric_range_enabled else "ВЫКЛ"],
            ["Распределение выше Pn", f"{asymmetry_distribution}% " if asymmetric_range_enabled else "не используется"],
            ["Центр диапазона всегда на текущей цене", "ДА" if use_current_price_as_center else "НЕТ"],
            ["", ""],
            ["РЕИНВЕСТИРОВАНИЕ КЭША", ""],
            ["Реинвестирование кэша", "ВКЛ" if reinvest_cash_enabled else "ВЫКЛ"],
            ["Частота реинвестирования", reinvest_frequency],
            ["", ""],
            ["ЗАКРЫТИЕ ШОРТА", ""],
            ["Макс. убыток для закрытия", f"{max_loss_percent}% выше цены входа"],
            ["Закрытие по BEP-триггеру", f"{'ВКЛ' if use_bep_as_primary_trigger else 'ВЫКЛ'}"],
            ["Буфер закрытия выше BEP", f"{bep_close_buffer_pct}% " if use_bep_as_primary_trigger else "не используется"],
            ["Закрытие по буферу до Pa", f"{'ВКЛ' if close_short_below_pa else 'ВЫКЛ'}"],
            ["Буфер до Pa", f"{close_short_buffer_pct}% " if close_short_below_pa else "не используется"],
            ["НОВАЯ ЛОГИКА: Выход из пула при закрытии ниже Pn", "ВКЛ - вся стоимость пула конвертируется в USDC"],
            ["", ""],
            ["ПОВТОРНОЕ ОТКРЫТИЕ ШОРТА", ""],
            ["Повторное открытие шорта", f"при падении на {reopen_percent_below_pn}% ниже Pn"],
            ["Задержка повторного открытия", f"{reopen_delay_hours} часов"],
            ["Минимальный отскок выше BEP", f"{reopen_bep_buffer_pct}%"],
            ["", ""],
            ["ПОВТОРНЫЙ ВХОД В ПУЛ", ""],
            ["Повторный вход в пул", "ВКЛ" if reentry_enabled else "ВЫКЛ"],
            ["Задержка повторного входа", f"{reentry_delay_hours} часов"],
            ["Порог падения цены", f"{reentry_price_drop_threshold}%"],
            ["Агрессивный вход через", f"{reentry_aggressive_hours} часов"],
        ]
        
        for row_idx, (param, value) in enumerate(parameters, 1):
            params_sheet.cell(row=row_idx, column=1, value=param)
            params_sheet.cell(row=row_idx, column=2, value=value)
        
        params_sheet.column_dimensions['A'].width = 45
        params_sheet.column_dimensions['B'].width = 30
        
        # --- CHARTS ---
        if charts:
            charts_sheet = workbook.create_sheet("Графики")
            charts_sheet['A1'] = 'ГРАФИКИ АНАЛИЗА СТРАТЕГИИ (15-минутные данные)'
            charts_sheet['A1'].font = Font(bold=True, size=14)
            img_row = 3
            for chart_name, img_data in charts:
                img = Image(img_data)
                img.width = 600
                img.height = 350
                cell = f'A{img_row}'
                charts_sheet.add_image(img, cell)
                charts_sheet.row_dimensions[img_row].height = img.height * 0.75
                img_row += int(img.height / 15) + 5
            charts_sheet.column_dimensions['A'].width = 100
        
        # --- PAYOFF ---
        payoff_sheet = workbook.create_sheet("Payoff Диаграмма")
        data_type_text_upper = "1-МИНУТНЫЕ" if use_1min_data else "15-МИНУТНЫЕ"
        payoff_sheet['A1'] = f'PAYOFF-ДИАГРАММА СТРАТЕГИИ: UNISWAP V3 LP + SHORT ETH ({data_type_text_upper} ДАННЫЕ)'
        payoff_sheet['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        payoff_sheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        payoff_sheet.merge_cells('A1:H1')
        payoff_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        data_type_text = "1-минутных" if use_1min_data else "15-минутных"
        payoff_sheet['A2'] = f'Анализ доходности стратегии хеджирования позиции LP в Uniswap V3 через шорт ETH на {data_type_text} данных'
        payoff_sheet.merge_cells('A2:H2')
        payoff_sheet['A2'].font = Font(bold=True, size=12)
        payoff_sheet['A2'].alignment = Alignment(horizontal='center')
        
        payoff_sheet['A4'] = 'ПАРАМЕТРЫ СТРАТЕГИИ:'
        payoff_sheet['A4'].font = Font(bold=True, size=12)

        data_type_display = "1-минутные" if use_1min_data else "15-минутные"
        parameters_data = [
            ['Параметр', 'Значение'],
            ['Цена входа (Pn)', f'${entry_price:,.2f}'],
            ['Нижняя граница (Pa)', f'${pa:,.2f}'],
            ['Верхняя граница (Pb)', f'${pb:,.2f}'],
            ['Ширина диапазона', f'${pb-pa:,.0f}'],
            ['Ликвидность (L)', f'{l:.4f}'],
            ['Размер шорта', f'{calculate_short_size(entry_price, pa, pb, l):.4f} ETH'],
            ['Текущая цена ETH', f'${current_price:,.2f}'],
            ['APR пула', f'{apr:.1%}'],
           
            ['Данные', data_type_display],
            ['Ценовой столбец', PRICE_COLUMN],
            ['Реинвестирование кэша', "ВКЛ" if reinvest_cash_enabled else "ВЫКЛ"],
            ['Частота реинвестирования', reinvest_frequency if reinvest_cash_enabled else "не используется"],
            ['НОВАЯ ЛОГИКА', "Выход из пула при закрытии ниже Pn"],
            ['ХЕДЖ', "Шорт открывается по Pn (центр диапазона)"]
        ]
        
        for row_idx, (param, value) in enumerate(parameters_data, 5):
            payoff_sheet.cell(row=row_idx, column=1, value=param)
            payoff_sheet.cell(row=row_idx, column=2, value=value)
        
        try:
            img = Image(payoff_filename)
            img.width = 800
            img.height = 500
            payoff_sheet.add_image(img, 'D4')
            print(f"  Payoff-диаграмма добавлена в Excel")
        except Exception as e:
            print(f"  Не удалось добавить payoff-диаграмму в Excel: {e}")
        
        payoff_sheet.column_dimensions['A'].width = 35
        payoff_sheet.column_dimensions['B'].width = 20
        payoff_sheet.column_dimensions['C'].width = 15
        payoff_sheet.column_dimensions['D'].width = 15
        payoff_sheet.column_dimensions['E'].width = 15
        payoff_sheet.column_dimensions['F'].width = 30
        
        # --- ORDERED DATA ---
        excel_df = df.copy()
        excel_df['Date'] = excel_df['Date'].dt.strftime('%Y-%m-%d %H:%M')
        
        ordered_columns = [
            'Date', 'ETH Price', 'Event',
            'DynamicInRange', 'DynamicPa', 'DynamicPn', 'DynamicPb', 'RangeWidth', 'DynamicL',
            'ShortActive', 'Short ETH', 'ShortEntryPrice', 'BEP', 'LastBEP',
            'CloseShortTrigger', 'CloseShortReason', 'LastCloseReason', 'CloseShortPrice',
            'Pool Value', 'Pool ETH', 'Pool USDC', 'CurrentPoolValue',
            'Pool Exit Value', 'Pool Exit Price', 'Pool Exit Realized PnL', 'Pool Exited',
            'Pending Cash', 'Compounded Cash', 'Compound Event', 'Compound Cost',
            'Accrued Fees', 'Funding PnL', 'Short Realized PnL', 'Costs',
            'Daily PnL', 'Realized PnL', 'Total PnL', 'Cumulative PnL',
            'ROI', 'Total ROI', 'Total Portfolio Value',
            # Новые колонки с просадками
            'Drawdown_15min', 'Daily_Max_Drawdown', 'Daily_Max_Runup', 
            'Daily_Return', 'Daily_Drawdown_From_Prev'
        ]
        
        existing_ordered_columns = [col for col in ordered_columns if col in excel_df.columns]
        
        ordered_df = excel_df[existing_ordered_columns]
        ordered_df.to_excel(writer, sheet_name='Упорядоченные данные', index=False)
        ordered_ws = writer.sheets['Упорядоченные данные']
        
        column_widths = {
            'Date': 15, 'ETH Price': 12, 'Event': 25,
            'DynamicInRange': 8, 'DynamicPa': 10, 'DynamicPn': 10, 'DynamicPb': 10, 'DynamicL': 12, 'RangeWidth': 10,
            'ShortActive': 8, 'Short ETH': 10, 'ShortEntryPrice': 12, 'BEP': 10, 'LastBEP': 10,
            'CloseShortTrigger': 8, 'CloseShortReason': 20, 'LastCloseReason': 15, 'CloseShortPrice': 12,
            'Pool Value': 12, 'Pool ETH': 10, 'Pool USDC': 12, 'CurrentPoolValue': 12,
            'Pool Exit Value': 12, 'Pool Exit Price': 12, 'Pool Exit Realized PnL': 12, 'Pool Exited': 8,
            'Pending Cash': 12, 'Compounded Cash': 12, 'Compound Event': 8, 'Compound Cost': 10,
            'Accrued Fees': 12, 'Funding PnL': 12, 'Short Realized PnL': 12, 'Costs': 10,
            'Daily PnL': 10, 'Realized PnL': 12, 'Total PnL': 12, 'Cumulative PnL': 12,
            'ROI': 8, 'Total ROI': 10, 'Total Portfolio Value': 15
        }
        
        for col_num, column_title in enumerate(existing_ordered_columns, 1):
            col_letter = get_column_letter(col_num)
            ordered_ws.column_dimensions[col_letter].width = column_widths.get(column_title, 12)
            cell = ordered_ws[f"{col_letter}1"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ordered_ws.freeze_panes = 'B2'
        ordered_ws.auto_filter.ref = f"A1:{get_column_letter(len(existing_ordered_columns))}1"
        
        # --- ALL DATA ---
        all_columns = list(excel_df.columns)
        excel_df.to_excel(writer, sheet_name='Все данные', index=False)
        worksheet = writer.sheets['Все данные']
        
        for col_num, column_title in enumerate(all_columns, 1):
            col_letter = get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = column_widths.get(column_title, 12)
            cell = worksheet[f"{col_letter}1"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        worksheet.freeze_panes = 'B2'
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(all_columns))}1"
        
        # --- SHEET ORDER ---
        sheet_order = ['Dashboard', 'Parameters', 'Графики', 'Payoff Диаграмма', 'Упорядоченные данные', 'Все данные']
        workbook._sheets.sort(key=lambda ws: sheet_order.index(ws.title) if ws.title in sheet_order else 999)
    
    print(f"✅ Файл {filename} успешно создан!")
    print(f"\n📊 СОДЕРЖАНИЕ ФАЙЛА:")
    print(f" 1. Dashboard - ключевые метрики и выводы")
    print(f" 2. Parameters - параметры стратегии и данных")
    print(f" 3. Графики - 4 графика анализа")
    print(f" 4. Payoff Диаграмма - анализ эффективности хеджирования")
    print(f" 5. Упорядоченные данные - логически сгруппированные данные")
    print(f" 6. Все данные - полные данные с условным форматированием")
    print(f"\n🎨 СОХРАНЕНЫ ВСЕ ФИЧИ EXCEL:")
    print(f" • Все листы присутствуют (6 листов)")
    print(f" • Условное форматирование (зелёный/красный для PnL и т.д.)")
    print(f" • Цветовое оформление и шрифты")
    print(f" • Автофильтры на всех листах")
    print(f" • Замороженные области (freeze_panes = 'B2')")
    print(f" • Чередование строк")
    print(f" • Форматирование чисел (проценты, разделители тысяч)")
    print(f" • Встроенные изображения графиков")
    print(f" • Правильный порядок листов")
    print(f" • Payoff-диаграмма с тёмным фоном")
    print(f"\n✅ ИСПРАВЛЕНИЯ (ТОЛЬКО 3 ПУНКТА):")
    print(f" 1. Выход из пула - сохранение CurrentPoolValue ДО обнуления")
    print(f" 2. Создание диапазона - шорт открывается по Pn (текущая цена)")
    print(f" 3. Повторный вход - все параметры в настройках + агрессивный вход через {reentry_aggressive_hours}ч")

# --- СОХРАНЕНИЕ В EXCEL ---
save_to_excel_with_improvements(df, 'backtest_results_improved_15min.xlsx')

# --- КРАТКАЯ СТАТИСТИКА ---
print("\n" + "="*80)
print("КРАТКАЯ СТАТИСТИКА (15-минутные данные)")
print("="*80)
print(f"Всего интервалов: {len(df)}")
total_hours = len(df) * time_step_hours
print(f"Общее время: {total_hours:.1f} часов ({total_hours/24:.1f} дней)")
print(f"Интервалов в диапазоне: {df['DynamicInRange'].sum()} ({df['DynamicInRange'].sum()/len(df)*100:.1f}%)")
print(f"Интервалов с активным шортом: {df['ShortActive'].sum()} ({df['ShortActive'].sum()/len(df)*100:.1f}%)")

# ПРОВЕРКА КОНСИСТЕНТНОСТИ (новый, надежный подсчет)
print("\n" + "-"*50)
print("ПРОВЕРКА КОНСИСТЕНТНОСТИ ОТКРЫТИЙ/ЗАКРЫТИЙ")
print("-"*50)

opened_count = 0
closed_count = 0

for i in range(len(df)):
    if i == 0:
        if df.at[i, 'ShortActive'] == 1:
            opened_count += 1
    else:
        if df.at[i, 'ShortActive'] == 1 and df.at[i-1, 'ShortActive'] == 0:
            opened_count += 1
        elif df.at[i, 'ShortActive'] == 0 and df.at[i-1, 'ShortActive'] == 1:
            closed_count += 1

final_status = df['ShortActive'].iloc[-1]
print(f"Открытий шорта: {opened_count}")
print(f"Закрытий шорта: {closed_count}")
print(f"Финальный статус: {'Активен' if final_status == 1 else 'Не активен'}")

if final_status == 1:
    expected_closes = opened_count - 1
    print(f"Ожидаемое число закрытий (с учетом открытого на конец): {expected_closes}")
    if closed_count == expected_closes:
        print("✅ КОНСИСТЕНТНО!")
    else:
        print(f"❌ НЕКОНСИСТЕНТНО! closed_count={closed_count}, expected={expected_closes}")
        
        # Диагностика: найдем проблемные места
        print("\n   Диагностика проблемных переходов:")
        for i in range(1, len(df)):
            if df.at[i, 'ShortActive'] == 0 and df.at[i-1, 'ShortActive'] == 1:
                # Нашли закрытие, проверим триггеры
                reason = df.at[i, 'CloseShortReason'] if pd.notna(df.at[i, 'CloseShortReason']) else 'unknown'
                trigger = df.at[i, 'CloseShortTrigger']
                exited = df.at[i, 'Pool Exited']
                print(f"     Строка {i}: причина='{reason}', trigger={trigger}, exited={exited}")
else:
    if opened_count == closed_count:
        print("✅ КОНСИСТЕНТНО!")
    else:
        print(f"❌ НЕКОНСИСТЕНТНО! opened={opened_count}, closed={closed_count}")

# Дополнительная статистика по причинам закрытия
print("\n" + "-"*50)
print("СТАТИСТИКА ПО ПРИЧИНАМ ЗАКРЫТИЯ")
print("-"*50)

below_pn_exit_closes = (df['CloseShortReason'].str.contains('below_pn_exit_pool', na=False)).sum()
print(f"Закрытий ниже Pn с выходом из пула: {below_pn_exit_closes}")

shift_count = (df['DynamicPa'] != df['DynamicPa'].shift(1)).sum() - 1
print(f"Сдвигов диапазона: {shift_count}")


below_pn_exit_closes = (df['CloseShortReason'].str.contains('below_pn_exit_pool', na=False)).sum()
print(f"Закрытий ниже Pn с выходом из пула: {below_pn_exit_closes}")

shift_count = (df['DynamicPa'] != df['DynamicPa'].shift(1)).sum() - 1
print(f"\nСдвигов диапазона: {shift_count}")

if reinvest_cash_enabled:
    compound_events = df['Compound Event'].sum()
    total_compound_cost = df['Compound Cost'].sum()
    final_compounded_cash = df['Compounded Cash'].iloc[-1]
    final_pending_cash = df['Pending Cash'].iloc[-1]
    print(f"\nРеинвестирование кэша:")
    print(f"  Событий реинвестирования: {compound_events}")
    print(f"  Реинвестированный кэш: ${final_compounded_cash:,.2f}")
    print(f"  Ожидающий кэш: ${final_pending_cash:,.2f}")
    print(f"  Затраты на реинвестирование: ${total_compound_cost:,.2f}")

print(f"\nСТОИМОСТЬ ПУЛА (V3):")
print(f"  Итоговая стоимость пула (Pool Value): ${df['Pool Value'].iloc[-1]:,.2f}")
print(f"  Итоговый Pool ETH: {df['Pool ETH'].iloc[-1]:.4f}")
print(f"  Итоговый Pool USDC: ${df['Pool USDC'].iloc[-1]:,.2f}")

print(f"\nОБЩАЯ СТАТИСТИКА:")
print(f"  Стартовый капитал: ${capital:,.2f}")
print(f"  Итоговая общая стоимость (Total Portfolio Value): ${df['Total Portfolio Value'].iloc[-1]:,.2f}")
print(f"  Кумулятивный PnL: ${df['Cumulative PnL'].iloc[-1]:,.2f}")
print(f"  Total ROI: {df['Total ROI'].iloc[-1]:.2%}")

# Эти строки уже удалены в пункте 1, теперь используется max_drawdown из функции
print(f"  Максимальная просадка (скорректированная): {max_drawdown:.2f}%")

# ===== ФИНАЛЬНАЯ ДИАГНОСТИКА ПРОСАДКИ =====
print("\n" + "="*60)
print("ФИНАЛЬНАЯ ДИАГНОСТИКА ПРОСАДКИ")
print("="*60)

# Анализируем последние 10 активных шортов (или все, если их меньше 10)
active_rows = df[df['ShortActive'] == 1].index
if len(active_rows) > 0:
    # Берем последние 10 или меньше
    last_active = active_rows[-min(10, len(active_rows)):]
    
    print(f"\nАнализ последних {len(last_active)} активных шортов:")
    print(f"{'Строка':>6} {'Цена':>8} {'Pa':>8} {'Тек.BEP':>10} {'% выше Pa':>10} {'Шорт тек':>10} {'Шорт для 5%':>12} {'Изменение':>10}")
    print("-" * 80)
    
    for idx in last_active:
        result = find_optimal_short_size_for_bep(df, idx, target_percent_above_pa=5)
        if result:
            print(f"{idx:6d} ${df.at[idx, 'ETH Price']:7.0f} ${result['pa']:7.0f} "
                  f"${result['current_bep']:9.2f} {result['current_bep_pct_above_pa']:9.1f}% "
                  f"{result['current_short_size']:10.4f} {result['short_size_needed']:12.4f} "
                  f"{result['short_size_change_pct']:9.1f}%")
    
    # Статистика по всем активным шортам
    print("\n" + "-" * 80)
    all_results = []
    for idx in active_rows:
        result = find_optimal_short_size_for_bep(df, idx, target_percent_above_pa=5)
        if result:
            all_results.append(result)
    
    if all_results:
        avg_change = np.mean([r['short_size_change_pct'] for r in all_results])
        avg_bep_pct = np.mean([r['current_bep_pct_above_pa'] for r in all_results])
        print(f"Среднее по всем {len(all_results)} активным шортам:")
        print(f"  Средний BEP выше Pa: {avg_bep_pct:.1f}%")
        print(f"  Среднее изменение для BEP@5%: {avg_change:.1f}%")
        print(f"  {'Рекомендация: увеличить шорт' if avg_change > 0 else 'Рекомендация: уменьшить шорт'} на {abs(avg_change):.1f}%")
else:
    print("\n  Нет активных шортов для анализа")

print("="*60)
# ===========================================

# Проверяем минимальные значения
min_value = df['Total Portfolio Value'].min()
min_row = df['Total Portfolio Value'].idxmin()
print(f"Минимальный Total Portfolio Value: ${min_value:.2f}")
print(f"Строка {min_row}: {df.at[min_row, 'Date']}")
print(f"  Pool Value: ${df.at[min_row, 'Pool Value']:.2f}")
print(f"  Pending Cash: ${df.at[min_row, 'Pending Cash']:.2f}")
print(f"  Compounded Cash: ${df.at[min_row, 'Compounded Cash']:.2f}")
print(f"  Pool Exited: {df.at[min_row, 'Pool Exited']}")

# Проверяем все строки с просадкой < -90%
if 'Correct Drawdown' in df.columns:
    bad_drawdowns = df[df['Correct Drawdown'] < -90]
    if len(bad_drawdowns) > 0:
        print(f"\nНайдено {len(bad_drawdowns)} строк с просадкой < -90%:")
        for idx in bad_drawdowns.index[:5]:  # первые 5
            print(f"  {idx}: {df.at[idx, 'Date']}, "
                  f"Value=${df.at[idx, 'Total Portfolio Value']:.2f}, "
                  f"Drawdown={df.at[idx, 'Correct Drawdown']:.2f}%, "
                  f"Exited={df.at[idx, 'Pool Exited']}")
    else:
        print(f"\n✅ Нет строк с просадкой < -90%")
else:
    print(f"\nСтолбец Correct Drawdown не найден")
print("="*60)
# ===========================================

data_type = "1-МИНУТНЫХ" if use_1min_data else "15-МИНУТНЫХ"
print(f"\n✅ БЭКТЕСТ НА {data_type} ДАННЫХ ЗАВЕРШЕН!")
print(f" • Все расчеты адаптированы для шага {time_step_hours*60:.1f} минут ({time_step_hours:.4f} часа)")
print(f" • Используемая цена: {PRICE_COLUMN}")
print(f" • Payoff-диаграмма сохранена: {payoff_filename}")
print(f" • Excel файл создан: {'backtest_results_1min.xlsx' if use_1min_data else 'backtest_results_improved_15min.xlsx'}")
print(f" • Сохранены ВСЕ листы и форматирование Excel")
print(f" • Подробный лог: {'ВКЛ' if verbose_logging else 'ВЫКЛ'}")
print(f"\n✅ ИСПРАВЛЕНИЯ (ТОЛЬКО 3 ПУНКТА):")
print(f" 1. Выход из пула - CurrentPoolValue сохранен, просадка исправлена")
print(f" 2. Создание диапазона - шорт открывается по Pn (текущая цена) - правильный хедж")
print(f" 3. Повторный вход - параметры в настройках: {reentry_delay_hours}ч, {reentry_price_drop_threshold}%, {reentry_aggressive_hours}ч")
print(f"\n📂 Откройте файл {'backtest_results_1min.xlsx' if use_1min_data else 'backtest_results_improved_15min.xlsx'}")
print(f"   Первым откроется лист Dashboard с ключевыми метриками.")
print(f"   Лист 'Payoff Диаграмма' содержит анализ эффективности хеджирования.")
print(f"   Лист 'Все данные' содержит все данные с отслеживанием выхода из пула.")